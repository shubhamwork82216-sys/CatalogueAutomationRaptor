# =========================================================
# RAPTOR QC SUITE + CATALOGUE AUTOMATION (FINAL MERGED)
# © Raptor Supplies | Created by Shubham Sisodia
# =========================================================

import streamlit as st
import pandas as pd
import numpy as np
import re
import ast
import tempfile
import os
import time
import io
import json
import random
import hashlib
import imagehash
import cv2
import easyocr
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from collections import Counter
from PIL import Image, ImageChops
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
import base64
from rapidfuzz import fuzz
from rapidfuzz import process as fuzz_process
from sklearn.feature_extraction.text import TfidfVectorizer

# ── selenium / undetected-chrome ──────────────────────────
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import undetected_chromedriver as uc
import uuid
import zipfile
import pytesseract
import torch
from urllib.parse import urlparse
from pathlib import Path
from transformers import BlipProcessor, BlipForQuestionAnswering

# =========================================================
# PAGE CONFIG
# =========================================================
st.set_page_config(page_title="Raptor QC Suite + Catalogue Automation", layout="wide")

# =========================================================
# HEADER / LOGO
# =========================================================
LOGO_PATH = os.path.join(os.path.dirname(__file__), "rpt-logo.svg")
if os.path.exists(LOGO_PATH):
    st.image(LOGO_PATH, width=140)

st.title("Raptor QC Suite + Catalogue Automation")

# =========================================================
# SIDEBAR
# =========================================================
tool = st.sidebar.radio(
    "Select Tool",
    [
        "📦 Catalogue Automation",
        "🖼️ Image QC Suite",
        "📊 Product Data QC",
        "🔐 A/B Testing QC (Admin)",
        "💰 MRO Price Sheet Mapper",
    ]
)

# =========================================================
# ADMIN GATE
# =========================================================
def admin_gate():
    pwd = st.sidebar.text_input("Admin Password", type="password")
    if pwd != "raptor_admin_2026":
        st.warning("Admin access required")
        st.stop()


# =========================================================
# TIME FORMATTING HELPER
# =========================================================
def fmt_time(seconds):
    """Format seconds into h m s string."""
    seconds = max(0, int(seconds))
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    if h > 0:
        return f"{h}h {m}m {s}s"
    elif m > 0:
        return f"{m}m {s}s"
    else:
        return f"{s}s"


# =========================================================
# SELENIUM HELPER (shared — undetected-chrome)
# =========================================================
_uc_driver = None

def _start_uc_driver():
    global _uc_driver
    if _uc_driver is None:
        opts = uc.ChromeOptions()
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-gpu")
        _uc_driver = uc.Chrome(version_main=145, options=opts, use_subprocess=True)

def get_image_with_selenium(url):
    global _uc_driver
    _start_uc_driver()
    try:
        _uc_driver.get(url)
        time.sleep(2)
        img_b64 = _uc_driver.execute_script("""
            const img = document.querySelector("img");
            if (!img) return null;
            return fetch(img.src)
                .then(r => r.blob())
                .then(b => new Promise(resolve => {
                    const reader = new FileReader();
                    reader.onloadend = () => resolve(reader.result);
                    reader.readAsDataURL(b);
                }));
        """)
        if not img_b64:
            return None
        img_bytes = base64.b64decode(img_b64.split(",")[1])
        return Image.open(BytesIO(img_bytes))
    except Exception as e:
        print("selenium error:", e)
        return None


# =========================================================
# ─────────────────────────────────────────────────────────
# TOOL 1 — CATALOGUE AUTOMATION
# ─────────────────────────────────────────────────────────
# =========================================================
if tool == "📦 Catalogue Automation":

    st.header("📦 Catalog Onboarding – Step-1, Discontinue & Step-2")

    # ── CONFIG ──────────────────────────────────────────
    SIMILARITY_THRESHOLD = 90

    def _brand_slug(df_or_series):
        """
        Extract brand name from a DataFrame (looks for 'Brand' or 'brand' column)
        or directly from a Series/scalar, and return a filesystem-safe slug.
        e.g.  "Milwaukee Tool"  →  "Milwaukee_Tool"
              "3M"              →  "3M"
        """
        try:
            if isinstance(df_or_series, pd.DataFrame):
                for col in ["Brand", "brand", "brand_name"]:
                    if col in df_or_series.columns:
                        val = df_or_series[col].dropna().astype(str).str.strip()
                        val = val[~val.str.lower().isin(["nan", "none", ""])]
                        if not val.empty:
                            brand = val.iloc[0]
                            break
                else:
                    return "Unknown"
            else:
                brand = str(df_or_series).strip()
            # Remove characters not safe in filenames
            brand = re.sub(r'[\\/:*?"<>|]', '', brand)
            # Collapse whitespace to underscore
            brand = re.sub(r'\s+', '_', brand).strip('_')
            return brand if brand and brand.lower() not in ("nan", "none") else "Unknown"
        except Exception:
            return "Unknown"

    TEMPLATE_HEADERS = [
        "S.No","brand_url","Brand","Mpn","Label","Grainger Sku","Entity Id",
        "L3 Name","L3 Url","Item Name","Parent_name","Product_Title",
        "Accessories","Kits/Components","Spare Parts","Product Detail",
        "product_features","Image Name","Image URL 1","Image URL 2",
        "Image URL 3","Image URL 4","Image URL 5","Video Url","Datasheet",
        "Shipping length(inch)","Shipping height(inch)","Shipping width(inch)",
        "weight(lb)","Quantity","Pack QTY","Price(usd)","Margin","Discount (%)",
        "Price Source","Country of Origin","UPC","Cross Reference",
        "Alternate Product","Discontinue/Remove","Brand Update","MPN Update",
        "MRO Product","Shipping Assumed","Weight Assumed","Assignee",
        "Task Type","Data_Source","Source_PD_URL","Doubt",
        "Remark_1","Remark_2","Remark_3","Remark_4","Remark_5",
        "Ref1","Ref2","Ref3","Ref4","Ref5"
    ]

    LIVE_TO_TEMPLATE = {
        "brand_url":"brand_url","brand_name":"Brand","mpn":"Mpn",
        "grainger_sku":"Grainger Sku","entity_id":"Entity Id",
        "L3_entity_name":"L3 Name","item_name":"Item Name",
        "parent_name":"Parent_name","title":"Product_Title",
        "image_name":"Image Name","Video_url":"Video Url",
        "datasheet":"Datasheet","ship_length":"Shipping length(inch)",
        "ship_height":"Shipping height(inch)","ship_width":"Shipping width(inch)",
        "weight":"weight(lb)","country_of_origin":"Country of Origin",
        "gtin":"UPC","cross_ref":"Cross Reference","mro":"Remark_5"
    }

    # NOTE: Description/Spec fields are now stored in Ref1-Ref5 (not via this map).
    # This map only drives the "From Scrape" column creation for key identity fields.
    SCRAPE_TO_TEMPLATE = {
        "Brand":"Brand","Mpn":"Mpn","Title":"Product_Title",
        "Video Links":"Video Url","Datasheets":"Datasheet",
    }
    # Source columns in Scrape file → Ref1-Ref5 destination columns
    REF_SCRAPE_MAP = {
        "Product Description 1": "Ref1",
        "Product Description 2": "Ref2",
        "Miscellaneous":         "Ref3",
        "Specification 1":       "Ref4",
        "Specification 2":       "Ref5",
    }

    # ── HELPERS ─────────────────────────────────────────
    def empty_template_df(rows):
        return pd.DataFrame("", index=range(rows), columns=TEMPLATE_HEADERS)

    def normalize_mpn(val):
        if pd.isna(val) or str(val).strip() == "":
            return ""
        val = str(val).strip().lower()
        if not val.startswith(("rp_", "rp_as")):
            return ""
        return re.sub(r"[^a-z0-9]", "", val)

    def safe_dict(val):
        try:
            return ast.literal_eval(val) if isinstance(val, str) else {}
        except:
            return {}

    def extract_images(val):
        urls = []
        if not val or str(val).strip() == "":
            return urls
        try:
            parsed = ast.literal_eval(val) if isinstance(val, str) else val
            if isinstance(parsed, list):
                for item in parsed:
                    if isinstance(item, dict) and "src" in item:
                        urls.append(str(item["src"]).strip())
                    elif isinstance(item, str):
                        urls.append(item.strip())
            elif isinstance(parsed, str) and parsed.startswith("http"):
                urls.append(parsed.strip())
        except:
            if isinstance(val, str) and val.startswith("http"):
                urls.append(val.strip())
        urls = [u for u in urls if u]
        return list(dict.fromkeys(urls))[:5]

    def extract_spec_fast(row):
        spec_combined = {}
        for col in ["Specification 1", "Specification 2", "Miscellaneous"]:
            spec_dict = safe_dict(row.get(col, ""))
            if isinstance(spec_dict, dict):
                spec_combined.update(spec_dict)
        out = {"l": [], "w": [], "h": [], "upc": "", "wt": []}
        for k, v in spec_combined.items():
            key = str(k).lower()
            val = str(v)
            if any(x in key for x in ["upc", "gtin", "ean", "barcode"]):
                digits = re.sub(r"\D", "", val)
                if digits and len(digits) >= 8:
                    out["upc"] = "rp_" + digits
            if "length" in key:
                out["l"].append(f"{k}: {v}")
            if "width" in key:
                out["w"].append(f"{k}: {v}")
            if "height" in key:
                out["h"].append(f"{k}: {v}")
            if "weight" in key:
                out["wt"].append(f"{k}: {v}")
        return {
            "Shipping length(inch)": "; ".join(out["l"]),
            "Shipping width(inch)":  "; ".join(out["w"]),
            "Shipping height(inch)": "; ".join(out["h"]),
            "UPC":                   out["upc"],
            "weight(lb)":            "; ".join(out["wt"]),
        }

    def crossref_match(live_row, scrape_df):
        live_mpn   = str(live_row.get("Clean_MPN", "")).strip()
        live_cross = str(live_row.get("cross_ref", "")).strip().lower()
        if not live_cross:
            return None
        for _, srow in scrape_df.iterrows():
            scrape_mpn   = str(srow.get("Clean_MPN", "")).strip()
            scrape_cross = str(srow.get("cross_ref", "")).strip().lower()
            if live_cross == scrape_mpn or scrape_cross == live_mpn:
                return srow
        return None

    # ── STEP-1 ──────────────────────────────────────────
    live_file  = st.file_uploader("Upload Live File",                    type="xlsx")
    scrape_file = st.file_uploader("Upload Scrape File",                 type="xlsx")
    disc_file  = st.file_uploader("Upload Discontinue File (optional)",  type="xlsx")

    if st.button("▶ Run Step-1 Mapping"):
        step1_start = time.time()
        step1_status = st.empty()
        step1_progress = st.progress(0)

        step1_status.info("⏳ Loading files…")
        live   = pd.read_excel(live_file)
        scrape = pd.read_excel(scrape_file)

        # ── Req 1: rp_ prefix validation ──────────────────────────────────────
        live_missing   = (~live["mpn"].astype(str).str.strip().str.lower()
                          .str.startswith(("rp_", "rp_as"))).sum()
        scrape_missing = (~scrape["Mpn"].astype(str).str.strip().str.lower()
                          .str.startswith(("rp_", "rp_as"))).sum()
        if live_missing > 0 or scrape_missing > 0:
            if live_missing > 0 and scrape_missing > 0:
                st.error(
                    f"❌ rp_ prefix missing in both Live Data ({live_missing} rows) "
                    f"and Scrape Data ({scrape_missing} rows). "
                    f"All MPNs must start with 'rp_' before mapping can proceed."
                )
            elif live_missing > 0:
                st.error(
                    f"❌ rp_ prefix missing in Live Data — {live_missing} MPN(s) do not "
                    f"start with 'rp_'. Fix the file and re-upload."
                )
            else:
                st.error(
                    f"❌ rp_ prefix missing in Scrape Data — {scrape_missing} MPN(s) do not "
                    f"start with 'rp_'. Fix the file and re-upload."
                )
            st.stop()
        # ── End validation ─────────────────────────────────────────────────────

        live["Clean_MPN"]   = live["mpn"].apply(normalize_mpn)
        scrape["Clean_MPN"] = scrape["Mpn"].apply(normalize_mpn)

        live   = live[live["Clean_MPN"]   != ""].reset_index(drop=True)
        scrape = scrape[scrape["Clean_MPN"] != ""].reset_index(drop=True)

        existing_df = empty_template_df(len(live))
        existing_df["Clean_MPN"] = live["Clean_MPN"].values
        existing_df["Mpn"]       = live["mpn"].values

        for src, tgt in SCRAPE_TO_TEMPLATE.items():
            if tgt in existing_df.columns:
                ins = existing_df.columns.get_loc(tgt) + 1
                col_name = f"{tgt} From Scrape"
                if col_name not in existing_df.columns:
                    existing_df.insert(ins, col_name, "")

        extra_cols = ["UPC","Shipping length(inch)","Shipping width(inch)","Shipping height(inch)","weight(lb)"]
        for col in extra_cols:
            if col in existing_df.columns:
                ins = existing_df.columns.get_loc(col) + 1
                new_col = f"{col} From Scrape"
                if new_col not in existing_df.columns:
                    existing_df.insert(ins, new_col, "")

        total_live = len(live)
        step1_status.info(f"⏳ Mapping {total_live} live rows…")

        for i, row in live.iterrows():
            matched_scrape_row = None
            direct_match = scrape[scrape["Clean_MPN"] == row["Clean_MPN"]]
            if not direct_match.empty:
                matched_scrape_row = direct_match.iloc[0]
                existing_df.at[i, "Remark_1"] = "mapped"
            else:
                matched_scrape_row = crossref_match(row, scrape)
                if matched_scrape_row is not None:
                    existing_df.at[i, "Remark_1"] = "mapped (via cross-ref)"

            if matched_scrape_row is not None:
                # Identity / media "From Scrape" columns
                for src, tgt in SCRAPE_TO_TEMPLATE.items():
                    col_name = f"{tgt} From Scrape"
                    if col_name in existing_df.columns:
                        existing_df.at[i, col_name] = matched_scrape_row.get(src, "")
                # Ref1-Ref5: unified reference columns
                for src_col, ref_col in REF_SCRAPE_MAP.items():
                    existing_df.at[i, ref_col] = str(matched_scrape_row.get(src_col, ""))
                for j, u in enumerate(extract_images(matched_scrape_row.get("Image Urls", ""))):
                    existing_df.at[i, f"Image URL {j+1}"] = u
                spec_data = extract_spec_fast(matched_scrape_row)
                for k, v in spec_data.items():
                    if k in existing_df.columns:
                        existing_df.at[i, k] = v
                    if f"{k} From Scrape" in existing_df.columns:
                        existing_df.at[i, f"{k} From Scrape"] = v
                existing_df.at[i, "Source_PD_URL"] = matched_scrape_row.get("URL", "")
                existing_df.at[i, "L3 From Scrape"] = matched_scrape_row.get("Breadscrumbs", "")

            # Time estimate
            elapsed = time.time() - step1_start
            done    = i + 1
            eta     = (elapsed / done) * (total_live - done) if done > 0 else 0
            step1_progress.progress(done / total_live)
            step1_status.info(f"⏳ Mapping rows… {done}/{total_live} | Elapsed: {fmt_time(elapsed)} | ETA: {fmt_time(eta)}")

        for src, tgt in LIVE_TO_TEMPLATE.items():
            if src in live.columns:
                existing_df[tgt] = live[src]

        # Drop redundant From Scrape columns (Req 2) — Clean_MPN dropped later after similarity check
        _cols_to_drop = ["Remark_3 From Scrape", "Remark_4 From Scrape"]
        existing_df.drop(columns=[c for c in _cols_to_drop if c in existing_df.columns],
                         inplace=True)

        # NEW ONBOARDING
        onboarding_src = scrape[~scrape["Clean_MPN"].isin(live["Clean_MPN"])].reset_index(drop=True)
        onboarding_df  = empty_template_df(len(onboarding_src))
        onboarding_df["Clean_MPN"] = onboarding_src["Clean_MPN"]
        onboarding_df["Mpn"]       = onboarding_src["Mpn"]

        total_ob = len(onboarding_src)
        ob_progress = st.progress(0)
        ob_status   = st.empty()
        ob_start    = time.time()

        for idx in range(total_ob):
            row = onboarding_src.loc[idx]
            # Identity / media fields via SCRAPE_TO_TEMPLATE
            for src, tgt in SCRAPE_TO_TEMPLATE.items():
                if src in onboarding_src.columns and tgt in onboarding_df.columns:
                    onboarding_df.at[idx, tgt] = row.get(src, "")
            # Ref1-Ref5: unified reference columns
            for src_col, ref_col in REF_SCRAPE_MAP.items():
                onboarding_df.at[idx, ref_col] = str(row.get(src_col, ""))
            for j, u in enumerate(extract_images(row.get("Image Urls", ""))):
                onboarding_df.at[idx, f"Image URL {j+1}"] = u
            spec_data = extract_spec_fast(row)
            for k, v in spec_data.items():
                if k in onboarding_df.columns:
                    onboarding_df.at[idx, k] = v
            onboarding_df.at[idx, "L3 Name"]      = row.get("Breadscrumbs", "")
            onboarding_df.at[idx, "Source_PD_URL"] = row.get("URL", "")

            elapsed = time.time() - ob_start
            done    = idx + 1
            eta     = (elapsed / done) * (total_ob - done) if done > 0 else 0
            ob_progress.progress(done / total_ob)
            ob_status.info(f"⏳ Onboarding rows… {done}/{total_ob} | Elapsed: {fmt_time(elapsed)} | ETA: {fmt_time(eta)}")

        ob_status.empty()
        # Clean_MPN is still needed below for discontinue + similarity checks — drop later

        # DISCONTINUE
        discontinue_df = pd.DataFrame()
        if disc_file:
            disc = pd.read_excel(disc_file)
            disc["Clean_MPN"] = disc["mpn"].apply(normalize_mpn)
            disc = disc[disc["Clean_MPN"] != ""].reset_index(drop=True)
            intersect = disc[disc["Clean_MPN"].isin(onboarding_df["Clean_MPN"])]
            discontinue_df = empty_template_df(len(intersect))
            discontinue_df["Mpn"]       = intersect["mpn"].values
            discontinue_df["Clean_MPN"] = intersect["Clean_MPN"].values

        # SIMILARITY
        similarity_rows = []
        existing_group  = {}
        for _, row in existing_df.iterrows():
            existing_group.setdefault(row["Clean_MPN"][:4], []).append(row)
        for _, new_row in onboarding_df.iterrows():
            clean_new = new_row["Clean_MPN"]
            orig_new  = new_row["Mpn"]
            block     = existing_group.get(clean_new[:4], [])
            for ex_row in block:
                score = fuzz.ratio(clean_new, ex_row["Clean_MPN"])
                if score >= 85:
                    similarity_rows.append([orig_new, ex_row["Mpn"], score])
        similarity_df = pd.DataFrame(similarity_rows, columns=["New_Onboard_MPN","Existing_MPN","Similarity_%"])

        # IMAGE URLS
        def extract_all_urls(df, mapped_only=False):
            urls = []
            for _, row in df.iterrows():
                if mapped_only and "mapped" not in str(row.get("Remark_1", "")):
                    continue
                for i in range(1, 6):
                    val = row.get(f"Image URL {i}", "")
                    if str(val).strip():
                        urls.append(val)
            return pd.DataFrame({"image_url": urls})

        existing_image_df   = extract_all_urls(existing_df, mapped_only=True)
        onboarding_image_df = extract_all_urls(onboarding_df)

        # Drop internal Clean_MPN column before export (not needed in output)
        for _df in [existing_df, onboarding_df]:
            _df.drop(columns=[c for c in ["Clean_MPN"] if c in _df.columns], inplace=True)

        # EXPORT
        output_step1 = BytesIO()
        with pd.ExcelWriter(output_step1, engine="openpyxl") as w:
            existing_df.to_excel(w, "Existing", index=False)
            onboarding_df.to_excel(w, "New_Onboarding", index=False)
            discontinue_df.to_excel(w, "Discontinued", index=False)
            similarity_df.to_excel(w, "MPN_Similarity_Check", index=False)

        image_output = BytesIO()
        with pd.ExcelWriter(image_output, engine="openpyxl") as w:
            existing_image_df.to_excel(w, "Existing_Image_URLs", index=False)
            onboarding_image_df.to_excel(w, "New_Onboarding_Image_URLs", index=False)

        st.session_state["step1_main_file"]  = output_step1.getvalue()
        st.session_state["step1_image_file"] = image_output.getvalue()
        st.session_state["step1_brand"]      = _brand_slug(live)

        total_time = time.time() - step1_start
        step1_status.empty()
        step1_progress.empty()
        st.success(f"✅ Step-1 completed in {fmt_time(total_time)}")

    if "step1_main_file" in st.session_state:
        _s1_brand = st.session_state.get("step1_brand", "Unknown")
        st.download_button(
            "⬇ Download Step-1 Output",
            st.session_state["step1_main_file"],
            f"Catalog_Onboarding_Step1_{_s1_brand}.xlsx",
        )
    if "step1_image_file" in st.session_state:
        _s1_brand = st.session_state.get("step1_brand", "Unknown")
        st.download_button(
            "⬇ Download Image URL Output",
            st.session_state["step1_image_file"],
            f"Image_URL_Output_{_s1_brand}.xlsx",
        )

    # =========================================================
    # STEP-2 — Title & Item Name Automation (Title_Creation logic)
    # =========================================================
    st.markdown("---")
    st.subheader("🧠 Step-2: Title & Item Name Automation")

    # ── Session state ─────────────────────────────────────
    for _k, _v in {"s2_logs":[],"s2_running":False,"s2_done":False,"s2_output":None,"s2_progress":0.0}.items():
        if _k not in st.session_state:
            st.session_state[_k] = _v

    # ── Title-creation constants ───────────────────────────
    TC_TITLE_DEFAULT_MAX  = 90
    TC_TITLE_HARD_MAX     = 120
    TC_TEMPERATURE        = 0
    TC_TOP_P              = 1
    TC_RANDOM_SEED        = 42
    TC_API_RETRIES        = 6
    TC_VALIDATION_RETRIES = 3

    # Both sheets now use the same Ref1-Ref4 structure; only L3 hint + title source differ
    TC_EXISTING_COLUMN_MAP = {
        "l3 from scrape":            "Category / Breadcrumb Hint",
        "product_title from scrape": "Reference Title. Do NOT copy verbatim. Prioritise specs and descriptions available",
        "brand":                     "Brand",
        "ref1":                      "Description 1",
        "ref2":                      "Description 2",
        "ref3":                      "Description 3",
        "ref4":                      "Specifications",
    }
    TC_NEW_COLUMN_MAP = {
        "l3 name":       "Category / Breadcrumb Hint",
        "product_title": "Reference Title. Do NOT copy verbatim. Prioritise specs and descriptions below",
        "brand":         "Brand",
        "ref1":          "Description 1",
        "ref2":          "Description 2",
        "ref3":          "Description 3",
        "ref4":          "Specifications",
    }
    TC_EXISTING_FILTER_COL   = "remark_1"
    TC_EXISTING_FILTER_VALUE = "mapped"
    # Discontinue check scans Ref1-Ref5 (and falls back to all-column scan in _tc_disc_flag_row)
    TC_DISC_CHECK_COLS = ["ref1", "ref2", "ref3", "ref4", "ref5"]

    # ── Configurable attribute exclusion keywords (Req 6a) ────────────────────
    TC_EXCLUSION_KEYWORDS_DEFAULT = (
        "brand\nstandard\nlicense\nupc\nwarranty\nseries\nurl\ngtin\nuan\n"
        "pack qty\npack quantity\nqty\nquantity\norigin country\ncountry of origin\n"
        "certifications\ncertification\nul standard\nul listed\nul rating\n"
        "sku\nean\nbarcode\nprice\nguarantee\ncompliance\ncompliant\nstandards"
    )
    # Persistent storage: keywords are saved to a file next to the script
    _EXCL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "exclusion_keywords.txt")

    def _excl_load():
        """Load keywords from file; fall back to hardcoded defaults if file missing."""
        try:
            with open(_EXCL_FILE, "r", encoding="utf-8") as f:
                return f.read()
        except FileNotFoundError:
            return TC_EXCLUSION_KEYWORDS_DEFAULT

    def _excl_save(text):
        """Save current keyword list to file."""
        with open(_EXCL_FILE, "w", encoding="utf-8") as f:
            f.write(text)

    # All keywords that flag a product as unavailable
    _TC_DISC_RE = re.compile(
        r"discontinu|not available|obsolete|obselet|obselete|"
        r"end of life|\beol\b|no longer available|phase.?out|phased out|superseded",
        re.IGNORECASE
    )

    def _tc_log(msg, kind="info"):
        tag_map = {"ok":"log-ok","warn":"log-warn","err":"log-err","info":"log-info","accent":"log-accent"}
        css = tag_map.get(kind, "log-info")
        st.session_state.s2_logs.append(f'<span class="{css}">{msg}</span>')

    def _tc_norm_cols(df):
        df = df.copy()
        df.columns = df.columns.str.strip().str.lower()
        return df  # never astype(str) — that turns blank cells into "nan"

    # Pre-compiled: strips dict KEY names  ('key': or "key":)  before searching values
    _DICT_KEY_RE = re.compile(r"""['"][^'"]*['"]\s*:""")

    def _tc_disc_check_parsed(obj):
        """
        Recursively check a parsed Python object for discontinue status.

        Rules (in priority order):
          1. If a dict contains 'discontinued_flag' key → return its boolean
             value directly. True = discontinued, False = not discontinued.
          2. Any key whose name contains 'discontinued' (e.g. 'discontinued_message')
             is IGNORED entirely — its value is never checked.
          3. All other string values are checked with _TC_DISC_RE
             (obsolete / no longer available / end of life / etc.).
          4. Lists → recurse into each element.
        """
        if obj is None or isinstance(obj, bool):
            return False
        if isinstance(obj, (int, float)):
            return False
        if isinstance(obj, str):
            return bool(_TC_DISC_RE.search(obj))
        if isinstance(obj, dict):
            # Priority 1: honour explicit discontinued_flag boolean
            for key, val in obj.items():
                if str(key).strip().lower() == "discontinued_flag":
                    return bool(val)   # True → disc, False → not disc
            # Priority 2 & 3: skip any key with 'discontinued' in name;
            # recurse into all other values
            for key, val in obj.items():
                if "discontinued" in str(key).strip().lower():
                    continue           # e.g. 'discontinued_message' — ignored
                if _tc_disc_check_parsed(val):
                    return True
            return False
        if isinstance(obj, (list, tuple)):
            return any(_tc_disc_check_parsed(item) for item in obj)
        return False

    def _tc_disc_in_value(v):
        """
        Return True if cell value v indicates the product is discontinued.

        • Dict/list cells (e.g. Grainger Ref3 metadata):
            → parsed with ast.literal_eval → _tc_disc_check_parsed (see above)
            → 'discontinued_flag': False  means NOT discontinued
            → 'discontinued_flag': True   means discontinued
            → keys like 'discontinued_message' are ignored completely
          If parsing fails, strip all key patterns then search remaining text.

        • Plain text cells (Ref1, Ref2, Ref4, Ref5):
            → _TC_DISC_RE checks for: discontinued / obsolete /
              no longer available / end of life / phased out / superseded
        """
        s = str(v).strip()
        if not s or s.lower() in ("nan", "none", ""):
            return False
        if s.startswith("{") or s.startswith("["):
            try:
                parsed = ast.literal_eval(s)
                return _tc_disc_check_parsed(parsed)
            except (ValueError, SyntaxError):
                # Parsing failed — strip key names, search remaining text only
                values_only = _DICT_KEY_RE.sub("", s)
                return bool(_TC_DISC_RE.search(values_only))
        return bool(_TC_DISC_RE.search(s))

    def _tc_is_discontinued(df_or_series, cols=None):
        """
        If df_or_series is a DataFrame, scan `cols` columns (OR logic across columns).
        If it's a Series, scan that series directly (legacy single-column use).

        Uses _tc_disc_in_value so dict-shaped cells (e.g. Ref3 with
        'discontinued_flag' as a key) are checked on VALUES only — the key
        name alone will NOT trigger a false positive.
        """
        if isinstance(df_or_series, pd.DataFrame):
            present = [c for c in (cols or TC_DISC_CHECK_COLS) if c in df_or_series.columns]
            if not present:
                return pd.Series(False, index=df_or_series.index)
            mask = pd.Series(False, index=df_or_series.index)
            for c in present:
                mask |= df_or_series[c].apply(_tc_disc_in_value)
            return mask
        # Legacy: single Series
        return df_or_series.apply(_tc_disc_in_value)

    def _tc_disc_flag_row(row):
        """Scan ALL columns of a row for discontinue keywords. Returns matched word or empty."""
        for col in row.index:
            if _tc_disc_in_value(row[col]):
                # Re-search the plain string to return the matched keyword text
                m = _TC_DISC_RE.search(str(row[col]))
                return m.group(0) if m else "discontinued"
        return ""

    # ── Title cleaning / validation helpers ───────────────
    def _tc_remove_series(text):
        if not isinstance(text, str): return text
        t = text
        for p in [r"\b[A-Za-z0-9]+(?:\s+[A-Za-z0-9]+){0,4}\s+Series\b",
                  r"\bSeries\s+[A-Za-z0-9]+(?:\s+[A-Za-z0-9]+){0,3}\b",
                  r"\bProduct\s+Line\b",r"\bProduct\s+Series\b",r"\bSeries\b"]:
            t = re.sub(p, "", t, flags=re.IGNORECASE)
        t = re.sub(r"\s*,\s*", ", ", t)
        t = re.sub(r",\s*,+", ", ", t)
        t = re.sub(r"\s{2,}", " ", t)
        return t.strip(" ,-/|")

    def _tc_remove_compliance(title):
        if not isinstance(title, str): return title
        parts = [p.strip() for p in title.split(",") if p.strip()]
        cw = re.compile(r"\b(compliant|compliance|certified|certification|standard|standards|certifications)\b", re.IGNORECASE)
        parts = [p for p in parts if not cw.search(p)]
        t = ", ".join(parts)
        t = re.sub(r"\s*,\s*", ", ", t)
        t = re.sub(r",\s*,+", ", ", t)
        t = re.sub(r"\s+", " ", t)
        return t.strip(" ,-/|")

    def _tc_normalize_units(title):
        if not isinstance(title, str): return title
        t = title
        subs = [
            (r"\b(\d+(?:\.\d+)?)\s*V\s*AC\b",    r"\1 Vac"),
            (r"\b(\d+(?:\.\d+)?)\s*VAC\b",        r"\1 Vac"),
            (r"\b(\d+(?:\.\d+)?)\s*V\s*DC\b",     r"\1 Vdc"),
            (r"\b(\d+(?:\.\d+)?)\s*VDC\b",        r"\1 Vdc"),
            (r"\b(\d+(?:\.\d+)?)\s*V\b",          r"\1 V"),
            (r"\b(\d+(?:\.\d+)?)\s*AMPS\b",       r"\1 Amp"),
            (r"\b(\d+(?:\.\d+)?)\s*AMP\b",        r"\1 Amp"),
            (r"\b(\d+(?:\.\d+)?)\s*A\b",          r"\1 Amp"),
            (r"\b(\d+(?:\.\d+)?)\s*KA\b",         r"\1 kA"),
            (r"\b(\d+(?:\.\d+)?)\s*KV\b",         r"\1 kV"),
            (r"\b(\d+(?:\.\d+)?)\s*HP\b",         r"\1 Hp"),
            (r"\b(\d+(?:\.\d+)?)\s*KW\b",         r"\1 kW"),
            (r"\b(\d+(?:\.\d+)?)\s*W\b",          r"\1 W"),
            (r"\b(\d+(?:\.\d+)?)\s*INCHES\b",     r"\1 Inch"),
            (r"\b(\d+(?:\.\d+)?)\s*INCH\b",       r"\1 Inch"),
            (r"\b(\d+(?:\.\d+)?)\s*IN\b",         r"\1 Inch"),
            (r"\b(\d+(?:\.\d+)?)\s*MM\b",         r"\1 mm"),
            (r"(\d+(?:\.\d+)?)\s*°",              r"\1 Degree"),
            (r"\b(\d+(?:\.\d+)?)\s*DEGREES\b",    r"\1 Degree"),
            (r"\b(\d+(?:\.\d+)?)\s*DEGREE\b",     r"\1 Degree"),
            (r"\b(\d+(?:\.\d+)?)\s*DEG\b",        r"\1 Degree"),
        ]
        for pat, rep in subs:
            t = re.sub(pat, rep, t, flags=re.IGNORECASE)
        return t

    def _tc_basic_clean(title):
        if not isinstance(title, str): return "ERROR: Invalid title"
        t = title.strip()
        if t.startswith("ERROR:"): return t
        t = _tc_remove_series(t)
        t = _tc_remove_compliance(t)
        t = _tc_normalize_units(t)
        # Preserve / and - : "1/3 Inch" must stay "1/3 Inch", "1-1/2" must stay "1-1/2"
        t = re.sub(r'[™$®`?\"\'=*@|°+:;\\{}\[\]<>~!#%^&]', "", t)
        for frag in ["\u00e2\u20ac\u017e","\u00c2","\u00c3\u2014","\u00e2\u20ac\u201c","\u00b5",
                     "\u00c3\u2030","\u0153","\u00e2\u20ac","\u201c","\u00a2","\u2018","\u201d",
                     "\u2122","\u00ae"]:
            t = t.replace(frag, "")
        for pattern in [
            r"\bPvt\.?\b",r"\bLtd\.?\b",r"\bLimited\b",r"\bMPN\b",r"\bModel\s*Number\b",
            r"\bPart\s*Number\b",r"\bPart\s*No\b",r"\bManufacturing\s*Number\b",r"\bUPC\b",r"\bSKU\b",
            r"\bPack\b",r"\bPk\b",r"\bPcs\b",r"\bQty\b",r"\bQuantity\b",r"\bPackage\b",
            r"\bWarranty\b",r"\bGuarantee\b",r"\bCompliant\b",r"\bCompliance\b",
            r"\bCertified\b",r"\bCertification\b",r"\bCertifications\b",r"\bStandard\b",r"\bStandards\b",
        ]:
            t = re.sub(pattern, "", t, flags=re.IGNORECASE)
        t = re.sub(r"https?://\S+", "", t, flags=re.IGNORECASE)
        t = re.sub(r"\bwww\.\S+", "", t, flags=re.IGNORECASE)
        t = re.sub(r"\b\S+\.(com|net|org|in|co|io|html|php|aspx)\S*\b", "", t, flags=re.IGNORECASE)
        t = re.sub(r"\s*,\s*", ", ", t)
        t = re.sub(r",\s*,+", ", ", t)
        t = re.sub(r"\s+", " ", t)
        t = re.sub(r"\b(\w+)(\s+\1\b)+", r"\1", t, flags=re.IGNORECASE)
        return t.strip(" ,-/|").strip()

    _TC_ITEM_END_WORDS = {
        "Sensor","Switch","Valve","Pump","Motor","Bearing","Gasket","Seal","Hose",
        "Fitting","Adapter","Adaptor","Connector","Cable","Relay","Fuse","Receptacle",
        "Plug","Socket","Mat","Wheel","Disc","Filter","Regulator","Cylinder","Gauge",
        "Transmitter","Controller","Bracket","Clamp","Washer","Bolt","Screw","Nut",
        "Tape","Light","Lamp","Module","Assembly","Kit","Tube","Pipe","Elbow","Tee",
        "Union","Coupling","Cap","Bushing","Sleeve","Rod","Plate","Block","Cover",
        "Handle","Lever","Spring","Nozzle","Caster","Thermostat","Contactor","Breaker",
        "Transformer","Actuator","Clutch","Brake","Manifold","Flange","Insert",
        "Cartridge","Solenoid","Coil","Strainer","Screen","Belt","Sheave","Pulley",
    }

    def _tc_enforce_comma(title):
        if not isinstance(title, str): return title
        t = _tc_basic_clean(title)
        if not t or t.startswith("ERROR:"): return t
        if "," in t: return t
        words = t.split()
        if len(words) <= 3: return t
        for i, word in enumerate(words[:6], start=1):
            cw = re.sub(r"[^A-Za-z]", "", word)
            if cw in _TC_ITEM_END_WORDS and i < len(words):
                return " ".join(words[:i]) + ", " + " ".join(words[i:])
        return " ".join(words[:2]) + ", " + " ".join(words[2:])

    def _tc_enforce_length(title):
        if not isinstance(title, str): return "ERROR: Invalid title"
        if title.startswith("ERROR:"): return title
        t = _tc_basic_clean(title)
        t = _tc_enforce_comma(t)
        if len(t) <= TC_TITLE_HARD_MAX: return t
        while len(t) > TC_TITLE_HARD_MAX and "," in t:
            parts = [p.strip() for p in t.split(",") if p.strip()]
            if len(parts) <= 2: break
            t = ", ".join(parts[:-1]).strip()
        if len(t) > TC_TITLE_HARD_MAX:
            t = t[:TC_TITLE_HARD_MAX].rsplit(" ", 1)[0].strip()
        return t.strip(" ,-/|")

    def _tc_remove_brand(title, batch_df=None, row_position=None):
        if not isinstance(title, str): return title
        t = title
        if batch_df is not None and row_position is not None and "brand" in batch_df.columns:
            try:
                brand = str(batch_df.iloc[row_position]["brand"]).strip()
                if brand and brand.lower() not in ["nan", "none"]:
                    t = re.sub(rf"\b{re.escape(brand)}\b", "", t, flags=re.IGNORECASE)
            except Exception:
                pass
        t = re.sub(r"\s*,\s*", ", ", t)
        t = re.sub(r",\s*,+", ", ", t)
        t = re.sub(r"\s{2,}", " ", t)
        return t.strip(" ,-/|")

    def _tc_validate_single(title, brand=""):
        t = _tc_enforce_length(title)
        if not isinstance(t, str) or not t.strip(): return False, "blank title", t
        if t.startswith("ERROR:"): return False, t, t
        if "," not in t: return False, "missing comma", t
        if t.endswith(","): return False, "title ends with comma", t
        if len(t) > TC_TITLE_HARD_MAX: return False, "title over 120 characters", t
        parts = [p.strip() for p in t.split(",") if p.strip()]
        if len(parts) < 2: return False, "title does not have item name and specification", t
        if len(parts[0].split()) > 4: return False, "item name is too long", t
        lower_t = t.lower()
        if re.search(r"\bseries\b|\bproduct line\b", lower_t, flags=re.IGNORECASE):
            return False, "contains Series or Product Line", t
        if brand and brand.lower() not in ["nan", "none"]:
            if re.search(rf"\b{re.escape(brand.strip())}\b", t, flags=re.IGNORECASE):
                return False, f"contains brand name: {brand.strip()}", t
        forbidden_words = [
            "pvt","ltd","limited","mpn","model number","part number","part no",
            "manufacturing number","upc","sku","series","product line","price",
            "warranty","guarantee","pack","pk","pcs","qty","quantity","package",
            "compliant","compliance","certified","certification","certifications",
            "standard","standards","https","www",
            "premium","best","high quality","heavy duty","durable","superior",
            "excellent","perfect","ideal","compatible",
        ]
        for word in forbidden_words:
            if re.search(rf"\b{re.escape(word)}\b", lower_t, flags=re.IGNORECASE):
                return False, f"contains forbidden word: {word}", t
        if re.search(r"https?://|www\.|\.(com|net|org|in|co)\b", lower_t, flags=re.IGNORECASE):
            return False, "contains URL-like text", t
        if re.search(r"[^A-Za-z0-9\s,./ \-]", t):
            return False, "contains special character", t
        if re.search(r"\b[A-Z]{2,6}[-_]?\d{2,}[A-Z0-9-]*\b", t):
            return False, "may contain model number or part number", t
        if re.search(r"\s{2,}", t):
            return False, "double spaces", t
        return True, "Valid", t

    def _tc_validate_titles(titles, batch_df=None):
        cleaned, statuses, invalid_reasons = [], [], []
        for i, title in enumerate(titles):
            brand = ""
            if batch_df is not None and "brand" in batch_df.columns:
                brand = str(batch_df.iloc[i]["brand"]).strip()
            t = _tc_remove_brand(title, batch_df, i)
            is_valid, reason, cleaned_t = _tc_validate_single(t, brand=brand)
            cleaned.append(cleaned_t)
            statuses.append(reason)
            if not is_valid:
                invalid_reasons.append(f"Product {i+1}: {reason} -> {cleaned_t}")
        return cleaned, statuses, invalid_reasons

    def _tc_extract_item_name(title):
        if not isinstance(title, str): return "ERROR: Invalid title"
        t = _tc_basic_clean(title)
        if not t: return "ERROR: Item name missing from title"
        if t.startswith("ERROR:"): return t
        if "," not in t: t = _tc_enforce_comma(t)
        item_name = t.split(",", 1)[0].strip()
        item_name = re.sub(r"\b\d+(\.\d+)?\b", "", item_name)
        item_name = _tc_remove_series(item_name)
        item_name = re.sub(r"\bSeries\b", "", item_name, flags=re.IGNORECASE)
        item_name = re.sub(r"\bProduct\s+Line\b", "", item_name, flags=re.IGNORECASE)
        item_name = re.sub(r"[^A-Za-z0-9\s]", "", item_name)
        item_name = re.sub(r"\s+", " ", item_name).strip().strip(" ,-/|")
        return item_name if item_name else "ERROR: Item name missing from title"

    def _tc_fuzzy_match(generated_name, reference_names, threshold=60):
        if not reference_names:
            return generated_name, None, "no_ref"
        if not generated_name or generated_name.startswith("ERROR:"):
            return generated_name, 0, "no_match"
        gen_clean = generated_name.strip().lower()
        gen_words = len(gen_clean.split())
        for ref in reference_names:
            if ref.strip().lower() == gen_clean:
                return ref, 100, "matched"
        same_length = [r for r in reference_names if len(r.strip().split()) == gen_words]
        if same_length:
            scored = [(r, fuzz.ratio(gen_clean, r.strip().lower())) for r in same_length]
            best_name, best_score = max(scored, key=lambda x: x[1])
            if best_score >= threshold:
                return best_name, round(best_score, 1), "matched"
        scored_all = [(r, fuzz.ratio(gen_clean, r.strip().lower())) for r in reference_names]
        best_name, best_score = max(scored_all, key=lambda x: x[1])
        best_score = round(best_score, 1)
        if best_score >= threshold:
            return best_name, best_score, "matched"
        return best_name, best_score, "low_confidence"

    def _tc_apply_fuzzy(item_names, reference_names, threshold=60):
        matched_names, scores, statuses = [], [], []
        for name in item_names:
            m, s, st = _tc_fuzzy_match(name, reference_names, threshold)
            matched_names.append(m); scores.append(s); statuses.append(st)
        return matched_names, scores, statuses

    def _tc_resolve_ref(match_str, ref_lower_map):
        """Case-insensitive O(1) lookup — returns canonical name or None."""
        return ref_lower_map.get(match_str.strip().lower())

    def _tc_suffix_reduce(nm, ref_lower_map):
        """
        Tier-2: Progressive suffix reduction — drops one word at a time from the
        front and checks for an exact match in the reference list.

        Examples:
          "Brass Plug Valve"      → tries "Plug Valve"  ✓ found → returns it
          "Long Stem Ball Valve"  → tries "Stem Ball Valve", "Ball Valve", "Valve"
          "Digital Tire Gauge"    → tries "Tire Gauge", "Gauge"
          "Temperature gauge"     → tries "gauge"  (if "Gauge" exists, found via lower)

        Returns (canonical_name, matched_suffix) or (None, None).
        """
        words = nm.strip().split()
        if len(words) <= 1:
            return None, None
        for drop in range(1, len(words)):
            suffix = " ".join(words[drop:])
            canon  = ref_lower_map.get(suffix.lower())
            if canon:
                return canon, suffix
        return None, None

    def _tc_top_candidates(nm, reference_names, ref_lower_map, top_k=30):
        """
        Pre-filter reference list to top_k most relevant candidates for item name `nm`,
        then FORCE-INCLUDE all suffix-generic names so AI always has fallback options.

        e.g. "Brass Plug Valve":
          scored top-30: Plug Valve, Ball Valve, Gate Valve, ...
          force-added  : "Plug Valve" (already in), "Valve" (added if missing)
        """
        nm_l   = nm.strip().lower()
        words  = nm_l.split()
        scored = sorted(
            [(r, fuzz.token_set_ratio(nm_l, r.strip().lower())) for r in reference_names],
            key=lambda x: x[1], reverse=True,
        )
        # Start with top-K by score
        candidates = [r for r, _ in scored[:top_k]]
        cand_set   = {c.strip().lower() for c in candidates}

        # Force-include every suffix generic that exists in reference list
        for drop in range(1, len(words)):
            suffix = " ".join(words[drop:])
            canon  = ref_lower_map.get(suffix.lower())
            if canon and canon.strip().lower() not in cand_set:
                candidates.append(canon)
                cand_set.add(canon.strip().lower())

        return candidates

    def _tc_ai_item_match_ctx(items_ctx, reference_names, ref_lower_map,
                               api_key, model_name, model_type="mistral",
                               batch_size=20, top_k=30):
        """
        Context-aware AI item matching with per-item pre-filtered candidate lists.

        How it works:
        1. For each item, pre-filter reference_names to top_k candidates via
           token_set_ratio — so the AI never sees 46,000+ names, only ~30 relevant ones.
        2. Send a batch of products to the AI. Each product block includes:
           - Item name, generated title, L3 category, Ref1-Ref4 descriptions
           - Its own CANDIDATES list (top 30 pre-filtered names)
        3. AI picks the best candidate per product based on full context.
        4. Result validated case-insensitively against reference list.

        Returns dict: {position_in_items_ctx (int) → canonical_ref_name (str)}
        """
        if not reference_names or not items_ctx:
            return {}

        result = {}

        for batch_start in range(0, len(items_ctx), batch_size):
            batch = items_ctx[batch_start: batch_start + batch_size]

            products_block = ""
            for i, ctx in enumerate(batch, start=1):
                nm         = ctx.get("name", "")
                candidates = _tc_top_candidates(nm, reference_names, ref_lower_map, top_k=top_k)
                cand_str   = "\n".join(
                    f"      {j+1}. {c}" for j, c in enumerate(candidates)
                )
                lines = [f"  Product {i}:"]
                lines.append(f"    Item Name : {nm}")
                if ctx.get("title"):
                    lines.append(f"    Title     : {ctx['title']}")
                if ctx.get("l3"):
                    lines.append(f"    Category  : {ctx['l3']}")
                for ri in range(1, 5):
                    v = ctx.get(f"ref{ri}", "")
                    if v and v.lower() not in ("nan", "none", ""):
                        lines.append(f"    Ref{ri}      : {v[:200]}")
                lines.append(f"    CANDIDATES:\n{cand_str}")
                products_block += "\n".join(lines) + "\n\n"

            prompt = (
                "You are a product taxonomy expert for an MRO catalogue.\n"
                "For EACH product, pick the SINGLE BEST name from its own CANDIDATES list.\n\n"
                "RULES:\n"
                "1. Read the full context (title, category, descriptions) to understand "
                "   what the product physically IS.\n"
                "2. Match semantically: 'Brass Plug Valve' → 'Plug Valve', "
                "   'Temperature gauge' → 'Temperature Gauge'.\n"
                "3. Prefer the most SPECIFIC candidate whose words are supported by the "
                "   context. If context mentions 'dual foot', prefer 'Dual Foot Tire Gauge' "
                "   over generic 'Gauge'.\n"
                "4. If context does NOT support a specific variant, pick the most appropriate "
                "   generic candidate (e.g. 'Gauge' rather than 'Truck Tire Gauge').\n"
                "5. NEVER pick a name from a completely different product category.\n"
                "6. Your answer MUST be copied EXACTLY from that product's CANDIDATES list.\n\n"
                f"PRODUCTS:\n{products_block}"
                "Strict JSON only. No markdown. No explanation.\n"
                '{"matches": [{"index": 1, "match": "exact name from candidates"}, ...]}'
            )

            try:
                if model_type == "gemini":
                    raw = _tc_call_gemini(prompt, api_key, model_name, retries=3)
                elif model_type == "ollama":
                    raw = _tc_call_ollama(prompt, model_name, retries=3)
                else:
                    raw = _tc_call_mistral(prompt, api_key, model_name, retries=3)

                clean = re.sub(r"^```(?:json)?", "", raw.strip()).strip().rstrip("`").strip()
                s, e  = clean.find("{"), clean.rfind("}")
                data  = json.loads(clean[s:e+1])

                for item in data.get("matches", []):
                    bidx  = int(item.get("index", 0)) - 1
                    match = str(item.get("match", "")).strip()
                    if 0 <= bidx < len(batch):
                        canon = _tc_resolve_ref(match, ref_lower_map)
                        if canon:
                            result[batch_start + bidx] = canon
            except Exception:
                pass  # batch left unmatched; caller writes no_match

        return result

    # Keys where ONLY the value is used — no attribute name appended
    _TC_VALUE_ONLY_KEYS = frozenset({
        "color", "colour", "finish", "material",
        "color/finish", "colour/finish", "finish/color", "color finish",
        "surface finish", "surface color",
    })

    def _tc_format_specs(raw_val, exclusion_set=None):
        """
        Parse a specification dict and return formatted strings:
          Color/Finish/Material  → value only          e.g. "Grey"
          Everything else        → "value AttrName"    e.g. "300V Voltage Rating"
        Skips keys matching the configurable exclusion_set.
        Preserves / and - exactly.
        """
        if not raw_val or str(raw_val).strip() in ("", "nan", "None", "NaN"):
            return []
        raw = str(raw_val).strip()
        if not raw.startswith("{"):
            return []
        import ast as _ast
        try:
            d = _ast.literal_eval(raw)
            if not isinstance(d, dict):
                return []
        except Exception:
            return []
        # Build effective exclusion set: defaults merged with any runtime additions
        _exc = exclusion_set if exclusion_set else set()
        result = []
        for k, v in d.items():
            k_str, v_str = str(k).strip(), str(v).strip()
            k_lower = k_str.lower()
            if not v_str or v_str.lower() in ("nan", "none", "n/a", "-", ""):
                continue
            # Skip if any exclusion keyword is contained in the attribute name
            if any(kw and kw in k_lower for kw in _exc):
                continue
            # Color / Finish / Material → value only
            if k_lower in _TC_VALUE_ONLY_KEYS or any(
                t in k_lower for t in ("color","colour","finish","material")
            ):
                result.append(v_str)
            else:
                result.append(f"{v_str} {k_str}")
        return result

    def _tc_build_prompt(batch_df, column_map, retry_feedback="", exclusion_set=None):
        products_block = ""
        for i, (_, row) in enumerate(batch_df.iterrows(), start=1):
            field_lines = []
            for col, role in column_map.items():
                if col not in row.index:
                    continue
                val = str(row[col]).strip()
                if val in ("", "nan", "None", "NaN"):
                    continue
                if "specification" in col.lower() or role.lower() == "specifications":
                    parts = _tc_format_specs(val, exclusion_set=exclusion_set)
                    if parts:
                        field_lines.append(f"    - {role}: " + " | ".join(parts))
                    elif not val.startswith("{"):
                        field_lines.append(f"    - {role}: {val}")
                else:
                    field_lines.append(f"    - {role}: {val}")
            products_block += f"\nProduct {i}:\n" + "\n".join(field_lines) + "\n"
        retry_block = ""
        if retry_feedback:
            retry_block = (
                f"\nIMPORTANT RETRY INSTRUCTION:\n"
                f"Previous output failed validation:\n{retry_feedback}\n"
                f"Fix ONLY the listed issues.\n"
            )
        return f"""You are a product listing expert for MRO e-commerce platforms.
Generate one product title per product listed below.

{products_block}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
TITLE FORMAT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Item Name, Spec1, Spec2, Spec3, Spec4

ITEM NAME: 1-4 words, singular, general — no specs, no brand, no symbols, no numbers.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
HOW TO USE SPECIFICATIONS — TWO STRICT RULES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Specs are pre-formatted as:  value AttributeName | value AttributeName | value

RULE 1 — Color, Finish, or Material specs are given as value only (no attr name after them).
  Use ONLY the value. Never add "Color", "Finish", or "Material" after it.
  ✓ Given: "Grey"        → use: Grey
  ✓ Given: "Steel"       → use: Steel
  ✗ WRONG: "Grey Color"        ← forbidden
  ✗ WRONG: "Steel Material"    ← forbidden

RULE 2 — All other specs are given as "value AttributeName" — copy BOTH into the title.
  ✓ Given: "300V Voltage Rating"    → use: 300V Voltage Rating
  ✓ Given: "22 AWG Wire Range"      → use: 22 AWG Wire Range
  ✓ Given: "105C Temp Rating"       → use: 105C Temp Rating
  ✓ Given: "3/8 Inch Size"          → use: 3/8 Inch Size
  ✓ Given: "1-1/2 NPT Thread Size"  → use: 1-1/2 NPT Thread Size
  ✗ WRONG: "300V"      ← attribute name dropped, forbidden
  ✗ WRONG: "22 AWG"    ← attribute name dropped, forbidden
  ✗ WRONG: "13 Inch"   ← fraction mangled, forbidden

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
OTHER RULES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- ONLY use specs listed for THIS product. Never invent values.
- Max 4 specs after item name. Pick most relevant for the product type.
- Always skip: Pack Qty, Origin Country, Certifications, UL, UPC, SKU, Warranty.
- No duplicate words or units.
- Length: default ≤ 90 chars, up to 120 only if needed, drop least important if over 120.
- Units: Vac, Vdc, V, Amp, kA, kV, Degree, Inch, mm, Hp, kW, W — space between value and unit.
- FORBIDDEN: Brand, MPN, Part number, Series, Pack, Qty, Warranty, Compliant,
  Certified, Standard, URLs, marketing adjectives (premium, heavy duty, best).

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
WORKED EXAMPLES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Specs: 300V Voltage Rating | 22 AWG Wire Range | 105C Temp Rating | Grey
Title: Wire Connector, 300V Voltage Rating, 22 AWG Wire Range, 105C Temp Rating, Grey

Specs: 90 Degree Angle | 3/8 Inch End | Carbon Steel
Title: Elbow, 90 Degree Angle, 3/8 Inch End, Carbon Steel

Specs: 12 To 24 Vdc Voltage Rating | Cylindrical M18 | Black
Title: Proximity Sensor, Cylindrical M18, 12 To 24 Vdc Voltage Rating, Black

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
RESPONSE FORMAT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Strict JSON only. No markdown. No explanation.
{{"titles": [{{"product_index": 1, "title": "..."}},{{"product_index": 2, "title": "..."}}]}}

{retry_block}"""

    def _tc_clean_json(text):
        text = str(text).strip()
        text = re.sub(r"^```(?:json)?", "", text).strip()
        text = re.sub(r"```$", "", text).strip()
        start, end = text.find("{"), text.rfind("}")
        if start != -1 and end != -1 and end > start:
            text = text[start:end+1]
        return text

    def _tc_parse_response(response_text, expected_count):
        data = json.loads(_tc_clean_json(response_text))
        titles = sorted(data.get("titles", []), key=lambda x: x.get("product_index", 0))
        result = [t.get("title", "ERROR: Missing title") for t in titles]
        while len(result) < expected_count:
            result.append("ERROR: Title missing from response")
        return result[:expected_count]

    def _tc_call_ollama(prompt, model_name, retries=TC_API_RETRIES):
        """Call a local Ollama model for title generation."""
        last_err = None
        for attempt in range(retries):
            try:
                r = requests.post(
                    "http://localhost:11434/api/generate",
                    json={"model": model_name, "prompt": prompt, "stream": False},
                    timeout=300,
                )
                r.raise_for_status()
                return r.json().get("response", "").strip()
            except Exception as e:
                last_err = e
                if attempt < retries - 1:
                    time.sleep(min(30, (2 ** attempt) + random.uniform(0.5, 1.5)))
        raise RuntimeError(f"Ollama failed after {retries} retries: {last_err}")

    def _tc_call_gemini(prompt, api_key, model_name, retries=TC_API_RETRIES):
        """Call Google Gemini API for title generation with exponential backoff."""
        last_err = None
        for attempt in range(retries):
            try:
                import google.generativeai as genai
                genai.configure(api_key=api_key)
                gm = genai.GenerativeModel(model_name)
                resp = gm.generate_content(
                    prompt,
                    generation_config=genai.types.GenerationConfig(temperature=0),
                )
                return resp.text.strip()
            except Exception as e:
                last_err = e
                msg = str(e).lower()
                retryable = any(x in msg for x in [
                    "429", "503", "quota", "rate", "timeout", "unavailable", "overloaded",
                ])
                if attempt < retries - 1 and retryable:
                    time.sleep(min(60, (2 ** attempt) + random.uniform(0.5, 2.0)))
                    continue
                raise last_err
        raise RuntimeError(f"Gemini failed after {retries} retries: {last_err}")

    def _tc_call_mistral(prompt, api_key, model_name, retries=TC_API_RETRIES):
        """Version-adaptive: tries mistralai v2 → v1 → v0 automatically."""
        def _extract(resp):
            c = resp.choices[0].message.content
            if isinstance(c, list):
                return "\n".join(
                    p.text if hasattr(p,"text") else p.get("text","")
                    for p in c if (hasattr(p,"text") and p.text) or (isinstance(p,dict) and p.get("text"))
                ).strip()
            return str(c).strip()

        def _call(p):
            try:  # v2.x
                from mistralai.client import Mistral as _M
                return _extract(_M(api_key=api_key,timeout_ms=120000).chat.complete(
                    model=model_name,messages=[{"role":"user","content":p}],
                    temperature=TC_TEMPERATURE,top_p=TC_TOP_P,random_seed=TC_RANDOM_SEED,
                    response_format={"type":"json_object"},timeout_ms=120000))
            except (ImportError,AttributeError): pass
            try:  # v1.x
                from mistralai import Mistral as _M
                return _extract(_M(api_key=api_key).chat.complete(
                    model=model_name,messages=[{"role":"user","content":p}],
                    temperature=TC_TEMPERATURE,top_p=TC_TOP_P,random_seed=TC_RANDOM_SEED,
                    response_format={"type":"json_object"}))
            except (ImportError,AttributeError): pass
            try:  # v0.x — uses plain dict messages to avoid unresolvable import
                from mistralai import MistralClient as _M
                _client_v0 = _M(api_key=api_key)
                # Build message as dict — avoids importing ChatMessage which
                # doesn't exist in v1/v2 and causes linter/IDE errors
                _msg = {"role": "user", "content": p}
                # Some v0 builds accept dicts, others need the object — try both
                try:
                    return _extract(_client_v0.chat(
                        model=model_name,
                        messages=[_msg],
                        temperature=TC_TEMPERATURE, top_p=TC_TOP_P,
                        random_seed=TC_RANDOM_SEED,
                        response_format={"type": "json_object"}))
                except TypeError:
                    # Strict v0 requires ChatMessage object — import inline
                    _cm = __import__(
                        "mistralai.models.chat_completion",
                        fromlist=["ChatMessage"]
                    ).ChatMessage
                    return _extract(_client_v0.chat(
                        model=model_name,
                        messages=[_cm(role="user", content=p)],
                        temperature=TC_TEMPERATURE, top_p=TC_TOP_P,
                        random_seed=TC_RANDOM_SEED,
                        response_format={"type": "json_object"}))
            except (ImportError, AttributeError): pass
            raise ImportError("No working mistralai SDK. Run: pip install --upgrade mistralai")

        last_error = None
        for attempt in range(retries):
            try:
                return _call(prompt)
            except Exception as e:
                last_error = e
                msg = str(e).lower()
                is_retryable = any(x in msg for x in [
                    "503","429","timeout","tempor","unavailable","rate limit",
                    "connection","overloaded","internal","timed out",
                ])
                if attempt < retries-1 and is_retryable:
                    time.sleep(min(60,(2**attempt)+random.uniform(0.5,2.0)))
                    continue
                raise last_error
        raise RuntimeError(f"Mistral failed after {retries} retries: {last_error}")

    def _tc_generate_batch(batch_df, column_map, api_key, model_name,
                           model_type="mistral", exclusion_set=None):
        """
        Smart retry: only failed rows are re-sent to the API each round.
        Passing rows are locked in immediately — zero wasted API calls.
        Supports model_type: 'mistral' | 'gemini' | 'ollama'
        """
        def _call_llm(prompt_txt):
            if model_type == "gemini":
                return _tc_call_gemini(prompt_txt, api_key, model_name, retries=3)
            elif model_type == "ollama":
                return _tc_call_ollama(prompt_txt, model_name, retries=3)
            else:  # default: mistral
                return _tc_call_mistral(prompt_txt, api_key, model_name, retries=3)

        n              = len(batch_df)
        final_titles   = ["ERROR: No title generated"] * n
        final_statuses = ["ERROR: No title generated"] * n
        pending        = list(range(n))

        for attempt in range(1, TC_VALIDATION_RETRIES + 1):
            if not pending:
                break
            sub_df = batch_df.iloc[pending].reset_index(drop=True)
            try:
                prompt  = _tc_build_prompt(
                    batch_df=sub_df, column_map=column_map,
                    retry_feedback="", exclusion_set=exclusion_set,
                )
                raw = _tc_parse_response(_call_llm(prompt), expected_count=len(sub_df))
                cleaned, statuses, invalid_reasons = _tc_validate_titles(raw, sub_df)
            except Exception as e:
                for pos in pending:
                    final_titles[pos]   = f"ERROR: {e}"
                    final_statuses[pos] = f"ERROR: {e}"
                break

            still_pending = []
            for sub_i, orig_pos in enumerate(pending):
                final_titles[orig_pos]   = cleaned[sub_i]
                final_statuses[orig_pos] = statuses[sub_i]
                if statuses[sub_i] not in ("Valid",) and not cleaned[sub_i].startswith("ERROR:"):
                    still_pending.append(orig_pos)
            pending = still_pending

        return final_titles, final_statuses

    def _tc_build_output_excel(df_existing_full, df_existing_filtered, existing_orig_idx,
                               df_new_full, df_new_filtered, new_orig_idx,
                               api_key, model_name, batch_size,
                               reference_names, fuzzy_threshold,
                               progress_callback, log_callback,
                               start_time_ref,
                               model_type="mistral",
                               exclusion_set=None,
                               scope=None):

        results = {}
        for label, df in [("Existing", df_existing_full), ("New_Onboarding", df_new_full)]:
            cols = list(df.columns)
            insert_pos = cols.index("catlv4") + 1 if "catlv4" in cols else len(cols)
            output_cols = (
                cols[:insert_pos]
                + ["standardized_item_name"]
                + cols[insert_pos:]
                + ["generated_title","title_validation_status",
                   "fuzzy_matched_item_name","fuzzy_score","fuzzy_match_status",
                   "discontinue_flag"]
            )
            out_df = df.reindex(columns=output_cols)
            for _nc in ["standardized_item_name","generated_title","title_validation_status",
                        "fuzzy_matched_item_name","fuzzy_score","fuzzy_match_status",
                        "discontinue_flag"]:
                if _nc in out_df.columns:
                    out_df[_nc] = out_df[_nc].astype(object)
            # Pre-fill discontinue_flag by scanning every column of every row
            for _idx in out_df.index:
                out_df.at[_idx, "discontinue_flag"] = _tc_disc_flag_row(out_df.loc[_idx])
            results[label] = out_df.copy()

        total_existing = len(df_existing_filtered)
        total_new      = len(df_new_filtered)
        total_rows     = total_existing + total_new
        processed      = [0]

        def _process_sheet(label, df_filtered, orig_idx, column_map):
            total = len(df_filtered)
            if total == 0:
                log_callback(f"⚠  No rows to process for [{label}]", "warn")
                return
            n_batches = (total + batch_size - 1) // batch_size
            log_callback(f"[{label}]  {total} rows · {n_batches} batches", "accent")

            all_item_names   = {}   # orig_idx → standardized_item_name
            all_titles       = {}   # orig_idx → generated_title
            all_statuses     = {}   # orig_idx → validation_status

            for b in range(n_batches):
                start    = b * batch_size
                end      = min(start + batch_size, total)
                batch_df = df_filtered.iloc[start:end]
                orig_idx_slice = list(orig_idx[start:end])

                elapsed     = time.time() - start_time_ref[0]
                done_so_far = processed[0]
                eta = (elapsed / done_so_far) * (total_rows - done_so_far) if done_so_far > 0 else 0
                log_callback(
                    f"  Batch {b+1}/{n_batches} — rows {start+1}–{end} | "
                    f"Elapsed: {fmt_time(elapsed)} | ETA: {fmt_time(eta)}", "info"
                )

                titles, statuses = _tc_generate_batch(
                    batch_df, column_map, api_key, model_name,
                    model_type=model_type, exclusion_set=exclusion_set,
                )
                item_names = [_tc_extract_item_name(t) for t in titles]

                out_df = results[label]
                for i, orig_i in enumerate(orig_idx_slice):
                    out_df.at[orig_i, "standardized_item_name"]  = item_names[i]
                    out_df.at[orig_i, "generated_title"]         = titles[i]
                    out_df.at[orig_i, "title_validation_status"] = statuses[i]
                    all_item_names[orig_i]  = item_names[i]
                    all_titles[orig_i]      = titles[i]
                    all_statuses[orig_i]    = statuses[i]

                processed[0] += len(batch_df)
                progress_callback(processed[0] / max(total_rows, 1))

                for i, (name, title, status) in enumerate(zip(item_names, titles, statuses)):
                    short_title = title[:55] + "…" if len(title) > 55 else title
                    log_callback(f"   Row {start+i+1}: {name} | {short_title} [{status}]", "ok")

                if b < n_batches - 1:
                    time.sleep(1.5)

            # ── Post-batch: UOM normalization pass (Req 6b) ──────────────────
            log_callback(f"[{label}] Running UOM normalization pass…", "info")
            out_df = results[label]
            for _idx in orig_idx:
                t = str(out_df.at[_idx, "generated_title"])
                if t and not t.startswith("ERROR:"):
                    out_df.at[_idx, "generated_title"] = _tc_normalize_units(t)

            # ── Post-batch: Item name matching (exact → AI with context+candidates) ─
            if reference_names:
                log_callback(f"[{label}] Running item name matching…", "info")

                # Build O(1) case-insensitive lookup dict once
                ref_lower_map = {r.strip().lower(): r for r in reference_names}

                # unique_ai_ctx  : one context dict per UNIQUE item name for AI
                # name_to_ai_pos : nm_lower → position in unique_ai_ctx
                # name_to_rows   : nm_lower → [all _idx rows with that name]
                unique_ai_ctx  = []
                name_to_ai_pos = {}
                name_to_rows   = {}

                for _idx in orig_idx:
                    nm = str(out_df.at[_idx, "standardized_item_name"]).strip()
                    if not nm or nm.startswith("ERROR:"):
                        out_df.at[_idx, "fuzzy_matched_item_name"] = ""
                        out_df.at[_idx, "fuzzy_score"]             = None
                        out_df.at[_idx, "fuzzy_match_status"]      = "no_item"
                        continue

                    # ── Tier 1: exact case-insensitive O(1) lookup ────────────
                    exact = ref_lower_map.get(nm.strip().lower())
                    if exact:
                        out_df.at[_idx, "fuzzy_matched_item_name"] = exact
                        out_df.at[_idx, "fuzzy_score"]             = 100
                        out_df.at[_idx, "fuzzy_match_status"]      = "exact_100"
                        continue

                    # ── Tier 2: progressive suffix reduction exact match ───────
                    suffix_match, _ = _tc_suffix_reduce(nm, ref_lower_map)
                    if suffix_match:
                        out_df.at[_idx, "fuzzy_matched_item_name"] = suffix_match
                        out_df.at[_idx, "fuzzy_score"]             = 100
                        out_df.at[_idx, "fuzzy_match_status"]      = "suffix_exact"
                        continue

                    # ── Tier 3: queue for AI — deduplicated by item name ───────
                    nm_key = nm.strip().lower()
                    name_to_rows.setdefault(nm_key, []).append(_idx)
                    if nm_key not in name_to_ai_pos:
                        # First occurrence → add context to AI queue
                        name_to_ai_pos[nm_key] = len(unique_ai_ctx)
                        row = out_df.loc[_idx]
                        unique_ai_ctx.append({
                            "name":  nm,
                            "title": str(row.get("generated_title", "")),
                            "l3":    str(row.get("l3 from scrape",
                                        row.get("l3 name",
                                        row.get("L3 From Scrape",
                                        row.get("L3 Name", ""))))),
                            "ref1":  str(row.get("ref1", row.get("Ref1", ""))),
                            "ref2":  str(row.get("ref2", row.get("Ref2", ""))),
                            "ref3":  str(row.get("ref3", row.get("Ref3", ""))),
                            "ref4":  str(row.get("ref4", row.get("Ref4", ""))),
                        })
                    # Duplicate rows skipped — result will be broadcast below

                # AI call — unique names only, pre-filtered candidates per item
                if unique_ai_ctx:
                    unique_count = len(unique_ai_ctx)
                    total_rows_ai = sum(len(v) for v in name_to_rows.values())
                    log_callback(
                        f"[{label}] AI context match: {unique_count} unique names "
                        f"covering {total_rows_ai} rows (top-30 candidates each)…", "info"
                    )
                    ai_map = _tc_ai_item_match_ctx(
                        unique_ai_ctx, reference_names, ref_lower_map,
                        api_key, model_name, model_type,
                        batch_size=20, top_k=30,
                    )
                    # Broadcast each AI result to ALL rows that share that item name
                    for nm_key, row_indices in name_to_rows.items():
                        pos      = name_to_ai_pos[nm_key]
                        ai_match = ai_map.get(pos, "")
                        for _idx in row_indices:
                            if ai_match:
                                out_df.at[_idx, "fuzzy_matched_item_name"] = ai_match
                                out_df.at[_idx, "fuzzy_score"]             = "AI"
                                out_df.at[_idx, "fuzzy_match_status"]      = "ai_semantic"
                            else:
                                nm_fb = str(out_df.at[_idx, "standardized_item_name"]).strip()
                                out_df.at[_idx, "fuzzy_matched_item_name"] = nm_fb
                                out_df.at[_idx, "fuzzy_score"]             = None
                                out_df.at[_idx, "fuzzy_match_status"]      = "no_match"

                log_callback(f"[{label}] ✅ Item name matching done", "ok")
            else:
                # No reference list — leave fuzzy columns blank
                for _idx in orig_idx:
                    out_df.at[_idx, "fuzzy_matched_item_name"] = ""
                    out_df.at[_idx, "fuzzy_score"]             = None
                    out_df.at[_idx, "fuzzy_match_status"]      = "no_ref"

        _active_scope = scope if scope else ["Existing", "New_Onboarding"]
        if "Existing" in _active_scope:
            _process_sheet("Existing", df_existing_filtered, existing_orig_idx, TC_EXISTING_COLUMN_MAP)
        else:
            log_callback("⏭  Existing skipped (not in scope)", "warn")
        if "New_Onboarding" in _active_scope:
            _process_sheet("New_Onboarding", df_new_filtered, new_orig_idx, TC_NEW_COLUMN_MAP)
        else:
            log_callback("⏭  New_Onboarding skipped (not in scope)", "warn")

        def _scrub_nan(df):
            out = df.copy()
            for col in out.select_dtypes(include="object").columns:
                out[col] = out[col].replace(
                    {"nan": "", "NaN": "", "None": "", "none": "", "NAN": ""},
                    regex=False
                )
            return out

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _scrub_nan(results["Existing"]).to_excel(
                writer, sheet_name="Existing", index=False)
            _scrub_nan(results["New_Onboarding"]).to_excel(
                writer, sheet_name="New_Onboarding", index=False)
        output.seek(0)
        return output.getvalue()

    # ── Step-2 UI ─────────────────────────────────────────
    step2_file = st.file_uploader("📂 Upload Step-1 Output File", type="xlsx", key="step2_file")

    # ── Model type ────────────────────────────────────────
    s2_model_type = st.radio(
        "🔧 Model Type",
        ["Mistral API", "Google Gemini API", "Local (Ollama)"],
        horizontal=True, key="s2_model_type",
    )

    s2_col_left, s2_col_right = st.columns(2, gap="large")
    with s2_col_left:
        if s2_model_type == "Mistral API":
            s2_api_key    = st.text_input("🔑 Mistral API Key", type="password", key="s2_api")
            s2_model_name = st.selectbox("🤖 Mistral Model",
                ["mistral-small-latest","mistral-medium-latest","mistral-large-latest","open-mistral-7b"],
                index=0, key="s2_model")
        elif s2_model_type == "Google Gemini API":
            s2_api_key    = st.text_input("🔑 Gemini API Key", type="password", key="s2_gemini_api")
            s2_model_name = st.selectbox("🤖 Gemini Model",
                ["gemini-2.0-flash","gemini-1.5-flash","gemini-1.5-pro","gemini-2.0-flash-lite"],
                index=0, key="s2_gemini_model")
        else:  # Ollama
            s2_api_key    = ""
            s2_model_name = st.selectbox("🤖 Ollama Model",
                ["qwen2.5","gemma3:4b","gemma3:12b","mistral","llama3.1"],
                index=0, key="s2_ollama_model")

        s2_batch_size = st.slider("📦 Batch Size", min_value=5, max_value=100, value=50, step=5, key="s2_batch")

    with s2_col_right:
        # Processing scope
        s2_scope = st.multiselect(
            "📋 Title Creation Scope",
            ["Existing", "New_Onboarding"],
            default=["Existing", "New_Onboarding"],
            key="s2_scope",
        )
        ref_file = st.file_uploader("📋 Item Name Reference (optional)", type=["xlsx","csv"], key="s2_ref")
        reference_names = []
        if ref_file:
            try:
                ref_df = pd.read_csv(ref_file) if ref_file.name.endswith(".csv") else pd.read_excel(ref_file)
                ref_file.seek(0)
                ref_col = st.selectbox("Column with item names", list(ref_df.columns), key="s2_ref_col")
                if ref_col:
                    reference_names = (
                        ref_df[ref_col].dropna().astype(str).str.strip()
                        .loc[lambda s: s.str.lower().isin(["nan","none"]) == False]
                        .tolist()
                    )
                    st.caption(f"✓ {len(reference_names)} names loaded")
            except Exception as e:
                st.error(f"Reference file error: {e}")

    # ── Attribute exclusion keywords (Req 6a) ────────────────────────────────
    # Load from file into session state on first run of this session
    if "s2_exclusion_raw" not in st.session_state:
        st.session_state["s2_exclusion_raw"] = _excl_load()

    with st.expander("🚫 Attribute Exclusion Keywords (one per line)", expanded=False):
        st.caption(
            "Spec attributes whose names **contain** any of these keywords are skipped "
            "during title generation. Click **💾 Save** to persist changes permanently "
            "across app restarts."
        )
        s2_exclusion_raw = st.text_area(
            "Exclusion Keywords",
            value=st.session_state["s2_exclusion_raw"],
            height=220,
            key="s2_exclusion_raw_input",
            label_visibility="collapsed",
        )
        # Update session state live as user types
        st.session_state["s2_exclusion_raw"] = s2_exclusion_raw

        if st.button("💾 Save Keywords to File", key="s2_excl_save"):
            try:
                _excl_save(s2_exclusion_raw)
                st.success(f"✅ Saved to {_EXCL_FILE}")
            except Exception as _ex:
                st.error(f"❌ Could not save: {_ex}")

    # Build the exclusion set at runtime from whatever is in the textarea
    s2_exclusion_set = {
        kw.strip().lower()
        for kw in s2_exclusion_raw.splitlines()
        if kw.strip()
    }

    _need_key = s2_model_type in ("Mistral API", "Google Gemini API")
    can_run_s2 = (
        step2_file is not None
        and (bool(s2_api_key) if _need_key else True)
        and bool(s2_scope)
        and not st.session_state.s2_running
    )

    if st.button("▶ Run Step-2 Title Automation", disabled=not can_run_s2):
        st.session_state.s2_running  = True
        st.session_state.s2_done     = False
        st.session_state.s2_logs     = []
        st.session_state.s2_output   = None
        st.session_state.s2_progress = 0.0
        st.rerun()

    if st.session_state.s2_running:
        s2_progress_bar = st.progress(st.session_state.s2_progress)
        s2_log_ph       = st.empty()

        def _render_logs():
            log_html = "<br>".join(st.session_state.s2_logs[-80:])
            s2_log_ph.markdown(f'<div style="background:#1a1a18;border-radius:8px;padding:1rem;font-family:monospace;font-size:0.7rem;color:#c8c7bc;max-height:300px;overflow-y:auto;line-height:1.8;">{log_html}</div>', unsafe_allow_html=True)

        def _update_progress(val):
            st.session_state.s2_progress = val
            s2_progress_bar.progress(min(val, 1.0))

        def _update_log(msg, kind="info"):
            _tc_log(msg, kind)
            _render_logs()

        try:
            step2_file.seek(0)
            xl = pd.ExcelFile(step2_file)
            sheet1_name, sheet2_name = xl.sheet_names[0], xl.sheet_names[1]

            step2_file.seek(0)
            df_existing = _tc_norm_cols(pd.read_excel(step2_file, sheet_name=sheet1_name, dtype=str))
            step2_file.seek(0)
            df_new      = _tc_norm_cols(pd.read_excel(step2_file, sheet_name=sheet2_name, dtype=str))

            _tc_log(f"Loaded Sheet 1: {len(df_existing)} rows", "ok")
            _tc_log(f"Loaded Sheet 2: {len(df_new)} rows", "ok")
            _render_logs()
            # Capture brand for filename
            st.session_state["s2_brand"] = _brand_slug(df_existing)

            # ── DEBUG: show every filter step so we can diagnose 0-eligible issues ──
            _rem1_col_exists = TC_EXISTING_FILTER_COL in df_existing.columns
            if _rem1_col_exists:
                mapped_mask = df_existing[TC_EXISTING_FILTER_COL].str.strip().str.lower() == TC_EXISTING_FILTER_VALUE.lower()
                _tc_log(f"  [DBG] Sheet1 '{TC_EXISTING_FILTER_COL}' col found ✔ | "
                        f"rows with '{TC_EXISTING_FILTER_VALUE}': {mapped_mask.sum()}", "info")
            else:
                mapped_mask = pd.Series(False, index=df_existing.index)
                _tc_log(f"  [DBG] Sheet1 column '{TC_EXISTING_FILTER_COL}' NOT FOUND — "
                        f"available cols: {list(df_existing.columns)}", "err")

            disc1_mask = _tc_is_discontinued(df_existing)   # scans all Ref cols
            _tc_log(f"  [DBG] Sheet1 disc-flagged rows: {disc1_mask.sum()} "
                    f"(of {len(df_existing)})", "info")
            # Show sample of any disc-flagged rows to help diagnose
            _disc1_samples = df_existing[disc1_mask].head(2)
            for _si, _srow in _disc1_samples.iterrows():
                for _sc in ["ref1","ref2","ref3","ref4","ref5"]:
                    if _sc in _srow and _tc_disc_in_value(_srow[_sc]):
                        _tc_log(f"    [DBG] Row {_si} flagged via {_sc}: "
                                f"{str(_srow[_sc])[:120]}", "warn")

            existing_mask = mapped_mask & ~disc1_mask
            df_existing_filtered = df_existing[existing_mask].copy()
            existing_orig_idx    = df_existing_filtered.index

            disc2_mask      = _tc_is_discontinued(df_new)      # scans all Ref cols
            _tc_log(f"  [DBG] Sheet2 disc-flagged rows: {disc2_mask.sum()} "
                    f"(of {len(df_new)})", "info")
            df_new_filtered = df_new[~disc2_mask].copy()
            new_orig_idx    = df_new_filtered.index

            _tc_log(f"Sheet 1: {len(df_existing_filtered)} eligible", "accent")
            _tc_log(f"Sheet 2: {len(df_new_filtered)} eligible", "accent")
            _render_logs()

            s2_start_ref = [time.time()]

            # Map UI model type string → internal key
            _s2_mt = {"Mistral API":"mistral","Google Gemini API":"gemini","Local (Ollama)":"ollama"}
            output_bytes = _tc_build_output_excel(
                df_existing_full=df_existing,
                df_existing_filtered=df_existing_filtered,
                existing_orig_idx=existing_orig_idx,
                df_new_full=df_new,
                df_new_filtered=df_new_filtered,
                new_orig_idx=new_orig_idx,
                api_key=s2_api_key,
                model_name=s2_model_name,
                batch_size=s2_batch_size,
                reference_names=reference_names,
                fuzzy_threshold=60,
                progress_callback=_update_progress,
                log_callback=_update_log,
                start_time_ref=s2_start_ref,
                model_type=_s2_mt.get(s2_model_type, "mistral"),
                exclusion_set=s2_exclusion_set,
                scope=s2_scope,
            )

            total_s2 = time.time() - s2_start_ref[0]
            st.session_state.s2_output  = output_bytes
            st.session_state.s2_done    = True
            st.session_state.s2_running = False
            _tc_log(f"Done in {fmt_time(total_s2)}. Download your output below.", "ok")
            _render_logs()
            _update_progress(1.0)
            st.rerun()

        except Exception as e:
            st.session_state.s2_running = False
            _tc_log(f"Fatal error: {e}", "err")
            _render_logs()
            st.error(f"Error: {e}")

    elif st.session_state.s2_logs:
        log_html = "<br>".join(st.session_state.s2_logs[-80:])
        st.markdown(f'<div style="background:#1a1a18;border-radius:8px;padding:1rem;font-family:monospace;font-size:0.7rem;color:#c8c7bc;max-height:300px;overflow-y:auto;line-height:1.8;">{log_html}</div>', unsafe_allow_html=True)

    if st.session_state.s2_done and st.session_state.s2_output:
        st.success("✅ Step-2 Completed")
        _s2_brand = st.session_state.get("s2_brand", "Unknown")
        st.download_button(
            "⬇ Download Step-2 Output",
            st.session_state.s2_output,
            f"Catalog_Onboarding_Step2_{_s2_brand}.xlsx",
        )

    # =========================================================
    # STEP-3 — L3 Mapping (Ollama)
    # =========================================================
    st.markdown("---")
    st.subheader("🧠 Step-3: L3 Mapping")

    step3_file = st.file_uploader("📂 Upload Step-2 Output File", type="xlsx", key="step3")
    l3_file    = st.file_uploader("📂 Upload L3 Master File",     type="xlsx", key="l3_file")

    # ── Engine toggle ─────────────────────────────────────────────────────────
    l3_engine = st.radio(
        "🔧 Mapping Engine",
        ["🖥️ Ollama (Local)", "☁️ Mistral API", "🌐 Google Gemini API"],
        horizontal=True, key="l3_engine",
        help="Ollama: free, local. Mistral/Gemini API: fast, needs API key."
    )
    use_mistral_api = l3_engine == "☁️ Mistral API"
    use_gemini_api  = l3_engine == "🌐 Google Gemini API"

    lc1, lc2, lc3, lc4, lc5 = st.columns(5)
    with lc1:
        if use_mistral_api:
            l3_mistral_key   = st.text_input("Mistral API Key", type="password", key="l3_mkey")
            l3_mistral_model = st.selectbox("Model",
                ["mistral-small-latest","mistral-medium-latest","mistral-large-latest"],
                key="l3_mmodel")
            l3_gemini_key = ""
        elif use_gemini_api:
            l3_gemini_key    = st.text_input("Gemini API Key", type="password", key="l3_gkey")
            l3_mistral_model = st.selectbox("Gemini Model",
                ["gemini-2.0-flash","gemini-1.5-flash","gemini-1.5-pro","gemini-2.0-flash-lite"],
                key="l3_gmodel")
            l3_mistral_key = ""
        else:
            model_choice     = st.selectbox("🤖 Ollama Model",
                ["qwen2.5","gemma3:4b","gemma3:12b","mistral","llama3.1"], key="l3_model")
            l3_mistral_key   = ""
            l3_gemini_key    = ""
            l3_mistral_model = ""
    with lc2:
        top_k = st.slider("🎯 TF-IDF Candidates", 10, 60, 25, key="l3_topk",
            help="L3 candidates shortlisted before LLM picks.")
    with lc3:
        batch_size = st.slider("📦 Batch Size", 1, 20, 5, key="l3_batch",
            help=(
                "Products per LLM call. "
                "Ollama: 3-5 recommended. "
                "API models: 5-10 recommended."
            ))
    with lc4:
        l3_url_col = st.text_input("L3 URL column", value="Url", key="l3_url_col",
            help="Column in L3 master with raptor URL e.g. /c/material-handling/c3/service-counters")
    with lc5:
        l3_rpm_limit = st.slider(
            "⏱ Max RPM", 10, 120, 60, 5, key="l3_rpm",
            help="Requests per minute cap for API models. Lower = safer against 429 errors.",
        ) if (use_mistral_api or use_gemini_api) else 999

    # Processing scope
    l3_scope = st.multiselect(
        "📋 L3 Assignment Scope",
        ["Existing", "New_Onboarding"],
        default=["Existing", "New_Onboarding"],
        key="l3_scope",
    )

    # Speed estimate
    _est_rows  = 149
    _est_call_s = 3 if (use_mistral_api or use_gemini_api) else 90
    _est_min   = (_est_rows / batch_size) * _est_call_s / 60
    st.caption(
        f"⚡ Estimated speed: batch={batch_size}, ~{_est_call_s}s/call → "
        f"**{_est_rows} rows ≈ {_est_min:.0f} min**"
    )

    def _l3_check_ollama():
        try: requests.get("http://localhost:11434", timeout=5); return True
        except: return False

    def _l3_check_model(m):
        try:
            r = requests.get("http://localhost:11434/api/tags", timeout=5)
            return any(m in x["name"] for x in r.json().get("models",[]))
        except: return False

    if st.button("▶ Run L3 Mapping", key="l3_run_btn"):
        if not step3_file:     st.error("❌ Upload Step-2 file"); st.stop()
        if not l3_file:        st.error("❌ Upload L3 Master file"); st.stop()
        if not l3_scope:       st.error("❌ Select at least one scope (Existing / New_Onboarding)"); st.stop()

        if use_mistral_api:
            if not l3_mistral_key:
                st.error("❌ Enter your Mistral API key"); st.stop()
        elif use_gemini_api:
            if not l3_gemini_key:
                st.error("❌ Enter your Gemini API key"); st.stop()
        else:
            if not _l3_check_ollama():
                st.error("❌ Ollama not running. Run: ollama serve"); st.stop()
            if not _l3_check_model(model_choice):
                st.error(f"❌ Model not pulled. Run: ollama pull {model_choice}"); st.stop()

        # ── Load data ─────────────────────────────────────────────────────────
        xl         = pd.ExcelFile(step3_file)
        df_exist   = xl.parse("Existing")
        df_onboard = xl.parse("New_Onboarding")
        _s3_brand  = _brand_slug(df_exist if len(df_exist) > 0 else df_onboard)
        l3_master  = pd.read_excel(l3_file)

        # ── Build L3 records from master ──────────────────────────────────────
        def _slug(s): return str(s).replace("-"," ").title() if s else ""
        def _parse_url(u):
            u = str(u).strip()
            m = re.search(r"/c/([^/]+)/c3/([^/?#]+)", u)
            if m: return _slug(m.group(1)), _slug(m.group(2))
            m2 = re.search(r"/c3/([^/?#]+)", u)
            if m2: return "", _slug(m2.group(1))
            return "", ""

        url_col_actual = l3_url_col if l3_url_col in l3_master.columns else next(
            (c for c in ("Url","url","URL","L3 Url","L3 URL") if c in l3_master.columns), None
        )
        l3_records, l3_names = {}, []
        for _, lr in l3_master.iterrows():
            nm = str(lr.get("L3 Name","")).strip()
            if not nm or nm.lower() in ("nan","none"): continue
            l1, l3s = _parse_url(lr.get(url_col_actual,"")) if url_col_actual else ("","")
            fp = f"{l1} > {nm}" if l1 else nm
            l3_records[nm] = {"l1":l1,"l3_slug":l3s,"full_path":fp,
                              "search_text":f"{nm} {l1} {l3s}".lower()}
            l3_names.append(nm)

        if not l3_names: st.error("❌ No L3 names found."); st.stop()
        st.info(f"📋 {len(l3_names)} L3 categories | URL col: {url_col_actual or 'not found'}")

        # ── TF-IDF index ──────────────────────────────────────────────────────
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.metrics.pairwise import cosine_similarity
        tfidf  = TfidfVectorizer(ngram_range=(1,2), min_df=1, sublinear_tf=True)
        l3_mat = tfidf.fit_transform([l3_records[n]["search_text"] for n in l3_names])
        st.success(f"✅ TF-IDF ready: {l3_mat.shape[0]} L3s")

        # ── Helpers ───────────────────────────────────────────────────────────
        def _norm(t):
            return " ".join(re.sub(r"[^a-z0-9 ]"," ",str(t).lower()).split())

        def _l1_from_bc(bc):
            bc = str(bc).strip()
            if not bc or bc.lower() in ("nan","none"): return ""
            return re.split(r"[>/|]", bc)[0].strip()

        def _shortlist(text, l1_hint="", k=top_k):
            q    = tfidf.transform([_norm(text+" "+l1_hint)])
            sims = cosine_similarity(q, l3_mat).flatten()
            if l1_hint:
                l1n = _norm(l1_hint)
                for ix, nm in enumerate(l3_names):
                    if l1n and l1n in _norm(l3_records[nm]["l1"]):
                        sims[ix] = min(1.0, sims[ix]+0.15)
            top = sims.argsort()[::-1][:k]
            return [l3_names[i] for i in top if sims[i]>0] or l3_names[:k]

        def _validate(pred, shortlist):
            pred = re.sub(r"[*`_#]","",str(pred))
            pred = re.sub(r"(?i)(the )?(best )?(matching )?(l3 )?"
                          r"(category|name|answer|option) *(is|:) *","",pred).strip().strip(".")
            if pred in l3_names:              return pred, "exact"
            for n in l3_names:
                if pred.lower()==n.lower():   return n, "exact"
            for n in shortlist:
                if pred.lower()==n.lower():   return n, "shortlist_exact"
            best_n, best_s = "", 0
            for n in l3_names:
                sc = fuzz.ratio(pred.lower(), n.lower())
                if sc > best_s: best_s, best_n = sc, n
            if best_s >= 88:                  return best_n, f"fuzzy_master_{best_s}"
            best_sn, best_ss = "", 0
            for n in shortlist:
                sc = fuzz.ratio(pred.lower(), n.lower())
                if sc > best_ss: best_ss, best_sn = sc, n
            if best_ss >= 75:                 return best_sn, f"fuzzy_shortlist_{best_ss}"
            for n in shortlist:
                if pred.lower() in n.lower() or n.lower() in pred.lower():
                    return n, "substring"
            return "REVIEW", "low_confidence"

        def _prod_text(row, bc_col):
            bc = ""
            for tc in [bc_col,"l3 from scrape","l3 name","breadcrumb","category",
                       "L3 From Scrape","L3 Name"]:
                v = str(row.get(tc, row.get(tc.title(),""))).strip()
                if v not in ("","nan","None"): bc = v; break
            # Gather all Ref columns that have content
            ref_lines = []
            for ri, rc in enumerate(["Ref1","Ref2","Ref3","Ref4","ref1","ref2","ref3","ref4"], 1):
                val = str(row.get(rc, "")).strip()
                if val and val.lower() not in ("nan","none",""):
                    ref_lines.append(f"REF{((ri-1)%4)+1}           : {val}")
                    if (ri % 4) == 0:
                        break   # already covered all 4 unique ref slots
            return "\n".join(filter(None, [
                "CATEGORY PATH  : " + str(bc),
                "GENERATED TITLE: " + str(row.get("generated_title",
                                         row.get("Generated Title",""))),
                "ITEM NAME      : " + str(row.get("standardized_item_name",
                                         row.get("Standardized Item Name",""))),
            ] + ref_lines))

        # ── LLM callers ───────────────────────────────────────────────────────
        def _ollama_call(prompt_text):
            try:
                r = requests.post("http://localhost:11434/api/generate",
                    json={"model": model_choice, "prompt": prompt_text, "stream": False},
                    timeout=300)
                r.raise_for_status()
                return r.json().get("response","").strip()
            except requests.exceptions.Timeout:
                return "REVIEW"
            except Exception as ex:
                st.warning(f"⚠️ Ollama error: {ex}"); return "REVIEW"

        def _mistral_call(prompt_text, retries=6):
            """Version-adaptive Mistral call with exponential-backoff retry on 429/503."""
            def _extract(resp):
                c = resp.choices[0].message.content
                return str(c).strip() if not isinstance(c, list) else " ".join(
                    p.text if hasattr(p,"text") else p.get("text","") for p in c).strip()

            def _call(p):
                try:
                    from mistralai.client import Mistral as _M
                    return _extract(_M(api_key=l3_mistral_key).chat.complete(
                        model=l3_mistral_model,
                        messages=[{"role":"user","content":p}],
                        temperature=0, max_tokens=500))
                except (ImportError, AttributeError): pass
                try:
                    from mistralai import Mistral as _M
                    return _extract(_M(api_key=l3_mistral_key).chat.complete(
                        model=l3_mistral_model,
                        messages=[{"role":"user","content":p}],
                        temperature=0, max_tokens=500))
                except (ImportError, AttributeError): pass
                raise ImportError("No working mistralai SDK found.")

            last_err = None
            for attempt in range(retries):
                try:
                    return _call(prompt_text)
                except Exception as e:
                    last_err = e
                    msg = str(e).lower()
                    retryable = any(x in msg for x in [
                        "429","503","rate limit","rate_limit","1300",
                        "timeout","unavailable","overloaded","connection",
                    ])
                    if attempt < retries - 1 and retryable:
                        wait = min(60, (2 ** attempt) + random.uniform(0.5, 2.0))
                        time.sleep(wait)
                        continue
                    raise last_err
            raise RuntimeError(f"Mistral failed after {retries} retries: {last_err}")

        def _gemini_call(prompt_text, retries=6):
            """Google Gemini API call with exponential-backoff retry."""
            last_err = None
            for attempt in range(retries):
                try:
                    import google.generativeai as genai
                    genai.configure(api_key=l3_gemini_key)
                    gm   = genai.GenerativeModel(l3_mistral_model)
                    resp = gm.generate_content(
                        prompt_text,
                        generation_config=genai.types.GenerationConfig(temperature=0),
                    )
                    return resp.text.strip()
                except Exception as e:
                    last_err = e
                    msg = str(e).lower()
                    retryable = any(x in msg for x in [
                        "429","503","quota","rate","timeout","unavailable","overloaded",
                    ])
                    if attempt < retries - 1 and retryable:
                        wait = min(60, (2 ** attempt) + random.uniform(0.5, 2.0))
                        time.sleep(wait)
                        continue
                    raise last_err
            raise RuntimeError(f"Gemini failed after {retries} retries: {last_err}")

        def _llm(prompt_text):
            if use_gemini_api:  return _gemini_call(prompt_text)
            if use_mistral_api: return _mistral_call(prompt_text)
            return _ollama_call(prompt_text)

        # ── RPM throttle state ────────────────────────────────────────────────
        _last_call_time = [0.0]
        _min_interval   = 60.0 / max(l3_rpm_limit, 1)   # seconds between API calls

        def _llm_throttled(prompt_text):
            """Wraps _llm with RPM-aware inter-call delay for API models."""
            if use_mistral_api or use_gemini_api:
                elapsed = time.time() - _last_call_time[0]
                gap     = _min_interval - elapsed
                if gap > 0:
                    time.sleep(gap)
            result = _llm(prompt_text)
            _last_call_time[0] = time.time()
            return result

        # ── Batch prompt builder ──────────────────────────────────────────────
        def _build_batch_prompt(items):
            """
            items: list of (prod_text, shortlist) tuples.
            Returns a prompt asking LLM to classify ALL products at once.
            Expected JSON response: {"results":[{"index":1,"l3":"..."},...]}
            """
            products_block = ""
            for i, (prod_text, shortlist) in enumerate(items, start=1):
                cats = "\n".join(
                    "  " + str(j+1) + ". " + str(l3) + "  [" + l3_records[l3]["full_path"] + "]"
                    for j, l3 in enumerate(shortlist)
                )
                products_block += (
                    "\n--- PRODUCT " + str(i) + " ---\n"
                    + prod_text + "\n"
                    + "CANDIDATES:\n" + cats + "\n"
                )
            return (
                "You are an expert MRO product classifier.\n"
                "Classify EACH product below to its best matching L3 category.\n\n"
                + products_block +
                "\nCLASSIFICATION RULES:\n"
                "1. READ all provided fields: title, item name, category path, and REF1-REF4 descriptions.\n"
                "2. UNDERSTAND the product's real-world APPLICATION and USE CASE — not just its name.\n"
                "   Example reasoning: A 'Handle' used for cabinet doors → 'Cabinet Accessories', "
                "   not a generic 'Handles' category.\n"
                "3. Use the FULL PATH in brackets to resolve ties between similarly named L3 categories.\n"
                "4. Pick EXACTLY ONE L3 from that product's own CANDIDATES list.\n"
                "5. Prefer the most specific applicable category. "
                "   Do NOT pick parent, accessory, or kit categories unless the product IS an accessory or kit.\n"
                "6. If the item name is generic (e.g. 'Fitting'), use the descriptions and category path "
                "   to determine the correct sub-category.\n\n"
                "RESPONSE FORMAT: Strict JSON only. No markdown. No explanation.\n"
                '{"results":[{"index":1,"l3":"exact L3 name"},{"index":2,"l3":"exact L3 name"}]}'
            )

        def _parse_batch_response(raw, expected_count, shortlists):
            """Parse JSON batch response, fall back to line-by-line if needed."""
            results = ["REVIEW"] * expected_count
            # Try JSON parse
            try:
                clean = re.sub(r"^```(?:json)?","",raw.strip()).strip().rstrip("`").strip()
                start, end = clean.find("{"), clean.rfind("}")
                if start != -1 and end > start:
                    data = json.loads(clean[start:end+1])
                    for item in data.get("results", []):
                        idx = int(item.get("index", 0)) - 1
                        if 0 <= idx < expected_count:
                            results[idx] = str(item.get("l3","REVIEW")).strip()
                    return results
            except Exception:
                pass
            # Fallback: split by lines, one L3 per line
            lines = [l.strip() for l in raw.strip().splitlines() if l.strip()]
            for i, line in enumerate(lines[:expected_count]):
                # Strip numbering like "1. Wire Connectors" or "Product 1: Wire Connectors"
                line = re.sub(r"^(product\s*)?\d+[\.\:\-\)]\s*","",line,flags=re.IGNORECASE)
                results[i] = line if line else "REVIEW"
            return results

        # ── Main mapping function (batched) ───────────────────────────────────
        def run_mapping(df, bc_col, sheet_label):
            df = df.copy()
            df["L3_Final"]      = ""
            df["l3_confidence"] = ""
            df["l3_l1_path"]    = ""

            cache      = {}  # (item_norm, l1_norm) → (l3, confidence)
            total      = len(df)
            llm_calls  = 0
            cache_hits = 0
            prog       = st.progress(0)
            status     = st.empty()
            t0         = time.time()

            rows_list  = list(df.iterrows())   # [(idx, row), ...]
            i          = 0

            while i < total:
                # ── Step 1: separate cache hits from rows needing LLM ─────────
                batch_llm_indices = []   # positions in rows_list needing LLM
                batch_llm_rows    = []   # (idx, row) for those

                for j in range(i, min(i + batch_size, total)):
                    idx, row = rows_list[j]
                    item_norm = _norm(str(row.get("standardized_item_name",
                                    row.get("Standardized Item Name",""))).strip())
                    bc_val = ""
                    for tc in [bc_col,"l3 from scrape","l3 name"]:
                        v = str(row.get(tc, row.get(tc.title(),""))).strip()
                        if v not in ("","nan","None"): bc_val = v; break
                    l1_norm = _norm(_l1_from_bc(bc_val))
                    cache_key = (item_norm, l1_norm)

                    if item_norm and cache_key in cache:
                        l3f, conf = cache[cache_key]
                        cache_hits += 1
                        df.at[idx, "L3_Final"]      = l3f
                        df.at[idx, "l3_confidence"] = conf + "_cached"
                        df.at[idx, "l3_l1_path"]    = l3_records.get(l3f,{}).get("full_path","")
                    else:
                        batch_llm_indices.append(j)
                        batch_llm_rows.append((idx, row, item_norm, l1_norm))

                # ── Step 2: batch LLM call for non-cached rows ────────────────
                if batch_llm_rows:
                    items_for_prompt = []
                    shortlists       = []
                    for idx, row, item_norm, l1_norm in batch_llm_rows:
                        prod_text = _prod_text(row, bc_col)
                        l1_hint   = str(row.get(bc_col, row.get(bc_col.title(),""))).strip()
                        sl        = _shortlist(prod_text, l1_hint=_l1_from_bc(l1_hint), k=top_k)
                        items_for_prompt.append((prod_text, sl))
                        shortlists.append(sl)

                    prompt_txt = _build_batch_prompt(items_for_prompt)
                    raw        = _llm_throttled(prompt_txt)
                    llm_calls += 1

                    preds = _parse_batch_response(raw, len(batch_llm_rows), shortlists)

                    for k2, (idx, row, item_norm, l1_norm) in enumerate(batch_llm_rows):
                        l3f, confidence = _validate(preds[k2], shortlists[k2])
                        cache_key = (item_norm, l1_norm)
                        if l3f != "REVIEW" and item_norm:
                            cache[cache_key] = (l3f, confidence)
                        df.at[idx, "L3_Final"]      = l3f
                        df.at[idx, "l3_confidence"] = confidence
                        df.at[idx, "l3_l1_path"]    = l3_records.get(l3f,{}).get("full_path","")

                # ── Progress update ───────────────────────────────────────────
                done    = min(i + batch_size, total)
                elapsed = time.time() - t0
                eta     = (elapsed / done) * (total - done) if done > 0 else 0
                last_l3 = df.at[rows_list[done-1][0], "L3_Final"]
                last_conf = df.at[rows_list[done-1][0], "l3_confidence"]
                prog.progress(done / total)
                status.info(
                    f"[{sheet_label}]  Rows {done}/{total}  |  "
                    f"LLM calls: {llm_calls}  |  "
                    f"Cache hits: {cache_hits}  |  "
                    f"Last: {last_l3} [{last_conf}]  |  "
                    f"Elapsed: {fmt_time(elapsed)}  |  ETA: {fmt_time(eta)}"
                )
                i += batch_size

            status.empty()
            return df

        # ── Run sheets based on scope ─────────────────────────────────────────
        # Same eligibility rules as Step 2:
        #   Existing     → remark_1 == "mapped"  AND  not discontinued
        #   New_Onboarding → not discontinued only
        # Non-eligible rows still appear in output; L3_Final is left blank for them.

        def _l3_eligible_mask(df, require_mapped):
            """Return boolean mask of rows eligible for L3 assignment."""
            norm = _tc_norm_cols(df)
            disc_mask = _tc_is_discontinued(norm)
            if require_mapped and "remark_1" in norm.columns:
                mapped_mask = norm["remark_1"].str.strip().str.lower() == "mapped"
            else:
                mapped_mask = pd.Series(True, index=norm.index)
            return mapped_mask & ~disc_mask

        def _run_sheet(df_full, bc_col, sheet_label, require_mapped):
            elig_mask = _l3_eligible_mask(df_full, require_mapped)
            n_elig    = elig_mask.sum()
            st.info(f"📋 [{sheet_label}]  {n_elig} eligible rows of {len(df_full)} → L3 mapping")

            # Pre-create output columns on the full df (blank for non-eligible)
            for col in ["L3_Final", "l3_confidence", "l3_l1_path"]:
                if col not in df_full.columns:
                    df_full[col] = ""

            if n_elig == 0:
                st.warning(f"⚠ [{sheet_label}] No eligible rows — skipping L3 mapping")
                return df_full

            df_elig   = df_full[elig_mask].copy()
            df_mapped = run_mapping(df_elig, bc_col, sheet_label)

            # Write results back to the full dataframe by index
            for col in ["L3_Final", "l3_confidence", "l3_l1_path"]:
                if col in df_mapped.columns:
                    df_full.loc[df_mapped.index, col] = df_mapped[col].values
            return df_full

        if "Existing" in l3_scope:
            with st.spinner("Mapping Existing sheet…"):
                df_exist = _run_sheet(df_exist, "L3 From Scrape", "Existing", require_mapped=True)
        else:
            st.info("⏭ Existing sheet skipped (not in scope)")

        if "New_Onboarding" in l3_scope:
            with st.spinner("Mapping New_Onboarding sheet…"):
                df_onboard = _run_sheet(df_onboard, "L3 Name", "New_Onboarding", require_mapped=False)
        else:
            st.info("⏭ New_Onboarding sheet skipped (not in scope)")

        # ── Summary stats ─────────────────────────────────────────────────────
        _sheet_map = []
        if "Existing"       in l3_scope: _sheet_map.append(("Existing",       df_exist))
        if "New_Onboarding" in l3_scope: _sheet_map.append(("New_Onboarding", df_onboard))

        for lbl, df in _sheet_map:
            t      = len(df)
            if "L3_Final" not in df.columns:
                st.info(f"📊 [{lbl}]  Skipped / no L3_Final column")
                continue
            review = (df["L3_Final"] == "REVIEW").sum()
            exact  = df["l3_confidence"].str.startswith("exact").sum()
            fm     = df["l3_confidence"].str.startswith("fuzzy_master").sum()
            fs     = df["l3_confidence"].str.startswith("fuzzy_shortlist").sum()
            cached = df["l3_confidence"].str.endswith("_cached").sum()
            st.info(
                f"📊 [{lbl}]  Total: {t}  |  Exact: {exact}  |  "
                f"Fuzzy(master): {fm}  |  Fuzzy(shortlist): {fs}  |  "
                f"Cached: {cached}  |  REVIEW: {review}"
            )

        # ── Export ────────────────────────────────────────────────────────────
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            df_exist.to_excel(w,   "Existing",       index=False)
            df_onboard.to_excel(w, "New_Onboarding", index=False)
        st.success("✅ Step-3 L3 Mapping Completed!")
        st.download_button(
            "⬇ Download Step-3 Output",
            out.getvalue(),
            f"Catalog_Onboarding_Step3_{_s3_brand}.xlsx",
        )

# =========================================================
# ─────────────────────────────────────────────────────────
# TOOL 2 — IMAGE QC SUITE  (New_Image_Resolution_Tool logic)
# ─────────────────────────────────────────────────────────
# =========================================================
elif tool == "🖼️ Image QC Suite":

    st.subheader("🖼️ Image QC Suite (Resolution | Reference | AI QC)")

    IQC_HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
        "Accept":     "image/webp,image/apng,image/*,*/*;q=0.8",
    }

    tab1, tab2, tab3, tab4 = st.tabs([
        "📐 Resolution Check",
        "🔍 Reference Image Check",
        "🧠 Advanced AI Image QC",
        "✨ SRGAN Image Enhancer"
    ])

    # ── TAB 1 — RESOLUTION CHECK ─────────────────────────
    with tab1:
        file1 = st.file_uploader("Upload Excel (image_url)", type="xlsx", key="r1")

        def analyze_resolution(url):
            try:
                r = requests.get(url, headers=IQC_HEADERS, timeout=15)
                if r.status_code == 200 and "image" in r.headers.get("Content-Type",""):
                    img = Image.open(BytesIO(r.content))
                elif r.status_code == 403:
                    img = get_image_with_selenium(url)
                else:
                    return url, None, None, None, None, f"{r.status_code}"
            except:
                img = get_image_with_selenium(url)

            if not img:
                return url, None, None, None, None, "Blocked"

            w, h = img.size
            gray    = np.array(img.convert("L"))
            bg_color = gray[0, 0]
            diff    = np.abs(gray.astype(int) - int(bg_color))
            mask    = diff > 15
            coords  = np.column_stack(np.where(mask))

            if len(coords) == 0:
                return url, w, h, w, h, "Blank Image"

            y_min, x_min = coords.min(axis=0)
            y_max, x_max = coords.max(axis=0)
            ow = x_max - x_min
            oh = y_max - y_min
            ratio = (ow * oh) / (w * h)
            pixel_density = mask.sum() / (w * h)

            if pixel_density < 0.05:
                remark = "CAD / Line Drawing"
            elif ratio < 0.3:
                remark = "Small object"
            else:
                remark = "Valid"

            return url, w, h, ow, oh, remark

        if file1 and st.button("Run Resolution Check", key="r1_btn"):
            xl           = pd.ExcelFile(file1)
            output_buffer = BytesIO()
            total_sheets = len(xl.sheet_names)
            sheet_prog   = st.progress(0)
            sheet_status = st.empty()
            global_start = time.time()

            with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
                for s_idx, sheet in enumerate(xl.sheet_names):
                    df = xl.parse(sheet)
                    if "image_url" not in df.columns:
                        continue

                    results     = []
                    total_rows  = len(df)
                    row_prog    = st.progress(0)
                    row_status  = st.empty()
                    sheet_start = time.time()

                    with ThreadPoolExecutor(max_workers=10) as ex:
                        futures = [ex.submit(analyze_resolution, u) for u in df["image_url"]]
                        for r_idx, f in enumerate(as_completed(futures)):
                            results.append(f.result())
                            done    = r_idx + 1
                            elapsed = time.time() - sheet_start
                            eta     = (elapsed / done) * (total_rows - done) if done > 0 else 0
                            row_prog.progress(done / total_rows)
                            row_status.info(f"Sheet [{sheet}] — {done}/{total_rows} | Elapsed: {fmt_time(elapsed)} | ETA: {fmt_time(eta)}")

                    row_status.empty()
                    out = pd.DataFrame(results, columns=["image_url","width","height","object_width","object_height","remark"])
                    out.to_excel(writer, sheet_name=f"{sheet}_Result"[:31], index=False)
                    sheet_prog.progress((s_idx+1) / total_sheets)
                    sheet_status.info(f"Sheet {s_idx+1}/{total_sheets} done")

            total_elapsed = time.time() - global_start
            st.success(f"✅ Resolution Check Completed in {fmt_time(total_elapsed)}")
            st.download_button("Download Excel", output_buffer.getvalue(), "resolution_multi.xlsx")

    # ── TAB 2 — REFERENCE IMAGE CHECK ────────────────────
    with tab2:
        file2 = st.file_uploader("Upload Excel (image_url)", type="xlsx", key="r2")
        ref   = st.text_input("Reference Image URL", key="ref_url")

        def get_image_bytes(url):
            try:
                r = requests.get(url, headers=IQC_HEADERS, timeout=15)
                if r.status_code == 200 and "image" in r.headers.get("Content-Type",""):
                    return r.content
                elif r.status_code == 403:
                    img = get_image_with_selenium(url)
                    if img:
                        buf = BytesIO()
                        img.save(buf, format="PNG")
                        return buf.getvalue()
            except:
                pass
            return None

        def _phash(data):
            try:
                return imagehash.phash(Image.open(BytesIO(data)).convert("RGB"))
            except:
                return None

        if file2 and ref and st.button("Run Reference Check", key="r2_btn"):
            xl           = pd.ExcelFile(file2)
            ref_bytes    = get_image_bytes(ref)
            ref_hash     = _phash(ref_bytes)
            output_buffer = BytesIO()
            total_sheets = len(xl.sheet_names)
            sheet_prog   = st.progress(0)
            global_start = time.time()

            with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
                for s_idx, sheet in enumerate(xl.sheet_names):
                    df = xl.parse(sheet)
                    if "image_url" not in df.columns:
                        continue
                    rows       = []
                    total_rows = len(df)
                    row_prog   = st.progress(0)
                    row_status = st.empty()
                    s_start    = time.time()

                    for r_idx, u in enumerate(df["image_url"]):
                        if str(u).strip() == ref.strip():
                            rows.append([u, "Matched"])
                        else:
                            b        = get_image_bytes(u)
                            img_hash = _phash(b) if b else None
                            match    = "Matched" if (img_hash and abs(img_hash - ref_hash) <= 5) else "Not Matched" if b else "Fetch Failed"
                            rows.append([u, match])

                        done    = r_idx + 1
                        elapsed = time.time() - s_start
                        eta     = (elapsed / done) * (total_rows - done) if done > 0 else 0
                        row_prog.progress(done / total_rows)
                        row_status.info(f"Sheet [{sheet}] — {done}/{total_rows} | Elapsed: {fmt_time(elapsed)} | ETA: {fmt_time(eta)}")

                    row_status.empty()
                    out = pd.DataFrame(rows, columns=["image_url","remark"])
                    out.to_excel(writer, sheet_name=f"{sheet}_Result"[:31], index=False)
                    sheet_prog.progress((s_idx+1) / total_sheets)

            total_elapsed = time.time() - global_start
            st.success(f"✅ Reference Check Completed in {fmt_time(total_elapsed)}")
            st.download_button("Download Excel", output_buffer.getvalue(), "reference_multi.xlsx")

    # ── TAB 3 — ADVANCED AI IMAGE QC (BLIP + OCR + Drawing) ──
    with tab3:

        st.markdown("### 🧠 Advanced AI Image QC")
        st.info(
            "Uses **Tesseract OCR** for text detection, **computer-vision heuristics** for "
            "watermark & drawing detection, and **BLIP VQA** (AI model) for watermark confirmation. "
            "⚠️ First run downloads ~1 GB BLIP model weights automatically."
        )

        file3 = st.file_uploader("Upload Excel (image_url column required)", type="xlsx", key="r3")

        # ── Config ──────────────────────────────────────────────────────────
        BLUR_THRESHOLD_QC = st.slider("Blur threshold (Laplacian variance)", 50, 300, 100, 10,
            help="Images with variance below this are flagged as blurry.", key="blur_thresh")

        c3l, c3r = st.columns(2)
        with c3l:
            workers3 = st.slider("Download threads", 1, 10, 5, key="r3_workers")
        with c3r:
            use_blip = st.checkbox(
                "Enable BLIP AI watermark check (slower, more accurate)",
                value=True, key="r3_blip",
                help="Disabling skips the AI model and uses only CV heuristics."
            )

        # ── BLIP model — cached so it loads once ────────────────────────────
        @st.cache_resource(show_spinner="Loading BLIP model (~1 GB, first run only)…")
        def _load_blip():
            proc  = BlipProcessor.from_pretrained("Salesforce/blip-vqa-base")
            model = BlipForQuestionAnswering.from_pretrained("Salesforce/blip-vqa-base")
            model.eval()
            return proc, model

        # ── Core analysis functions ─────────────────────────────────────────
        def _load_img_cv2(url):
            """Download URL and return BGR numpy array, or None on failure."""
            try:
                headers = {"User-Agent": "Mozilla/5.0"}
                resp = requests.get(url, timeout=15, headers=headers)
                arr  = np.frombuffer(resp.content, np.uint8)
                img  = cv2.imdecode(arr, cv2.IMREAD_COLOR)
                return img
            except:
                return None

        def _blur_check_cv2(img_cv2, threshold):
            gray = cv2.cvtColor(img_cv2, cv2.COLOR_BGR2GRAY)
            var  = cv2.Laplacian(gray, cv2.CV_64F).var()
            return round(var, 2), var < threshold

        def _ocr_text_present(img_cv2):
            """
            Tesseract OCR — only counts real words (len>=3, conf>60).
            Returns True if any such word is found.
            """
            try:
                gray = cv2.cvtColor(img_cv2, cv2.COLOR_BGR2GRAY)
                h, w = gray.shape
                if w < 1000:
                    scale = 1000 / w
                    gray  = cv2.resize(gray, None, fx=scale, fy=scale,
                                       interpolation=cv2.INTER_CUBIC)
                thresh = cv2.adaptiveThreshold(
                    gray, 255,
                    cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                    cv2.THRESH_BINARY, 31, 10
                )
                data = pytesseract.image_to_data(
                    thresh, config="--psm 11 --oem 3",
                    output_type=pytesseract.Output.DICT
                )
                real_words = [
                    t for t, c in zip(data["text"], data["conf"])
                    if str(t).strip() and int(c) > 60 and len(str(t).strip()) >= 3
                ]
                return len(real_words) > 0
            except Exception:
                return False

        def _blip_watermark(img_cv2, proc, model):
            """Ask BLIP: 'Is there a watermark?' — returns True/False."""
            try:
                from PIL import Image as PILImage
                image_pil = PILImage.fromarray(cv2.cvtColor(img_cv2, cv2.COLOR_BGR2RGB))
                inputs = proc(
                    image_pil,
                    "Is there a watermark on this image? Answer yes or no.",
                    return_tensors="pt"
                )
                with torch.no_grad():
                    out = model.generate(**inputs, max_new_tokens=10)
                answer = proc.decode(out[0], skip_special_tokens=True).strip().lower()
                return any(p in answer for p in ["yes", "true", "watermark", "it does"])
            except Exception:
                return False

        def _cv_watermark_check(img_cv2, use_blip_flag, blip_proc, blip_model):
            """
            Stage-1: CV heuristic (mid-tone contour density + saturation).
            If Stage-1 passes AND BLIP is enabled → confirm with BLIP.
            """
            gray = cv2.cvtColor(img_cv2, cv2.COLOR_BGR2GRAY)
            hsv  = cv2.cvtColor(img_cv2, cv2.COLOR_BGR2HSV)
            h, w = gray.shape
            sat_mean = float(hsv[:, :, 1].mean())

            mid_mask  = ((gray >= 80) & (gray <= 200)).astype(np.uint8) * 255
            kernel    = np.ones((3, 3), np.uint8)
            mid_clean = cv2.erode(mid_mask, kernel, iterations=1)
            contours, _ = cv2.findContours(mid_clean, cv2.RETR_EXTERNAL,
                                           cv2.CHAIN_APPROX_SIMPLE)
            text_like = 0
            for cnt in contours:
                area = cv2.contourArea(cnt)
                if area < 10:
                    continue
                _, _, cw, ch = cv2.boundingRect(cnt)
                aspect = max(cw, ch) / (min(cw, ch) + 1e-5)
                if cw > w * 0.3 or ch > h * 0.3:
                    continue
                if 20 < area < 2000 and aspect > 1.5:
                    text_like += 1

            density = text_like / (h * w / 10000)
            if density > 0.3 and sat_mean > 35:
                if use_blip_flag:
                    return _blip_watermark(img_cv2, blip_proc, blip_model)
                return True   # CV-only positive
            return False

        def _drawing_check(img_cv2):
            """
            Technical drawing detection:
            white_ratio>0.85, sat_mean<15, content_std>70, blur_score>500
            """
            gray = cv2.cvtColor(img_cv2, cv2.COLOR_BGR2GRAY)
            hsv  = cv2.cvtColor(img_cv2, cv2.COLOR_BGR2HSV)
            h, w = gray.shape
            white_ratio  = np.sum(gray > 235) / (h * w)
            sat_mean     = float(hsv[:, :, 1].mean())
            blur_score   = cv2.Laplacian(gray, cv2.CV_64F).var()
            non_white    = gray[gray <= 235]
            content_std  = float(np.std(non_white)) if len(non_white) > 100 else 0
            return (
                white_ratio > 0.85 and
                sat_mean < 15 and
                content_std > 70 and
                blur_score > 500
            )

        def _analyze_single(url, blur_thr, use_blip_flag, blip_proc, blip_model):
            img = _load_img_cv2(url)
            if img is None:
                return {
                    "url": url, "fetch_status": "Failed",
                    "blur_score": None, "blurry": None,
                    "text_detected": None, "watermark": None,
                    "autocad_drawing": None, "ai_remark": "Fetch failed"
                }
            blur_score, blurry  = _blur_check_cv2(img, blur_thr)
            text_present        = _ocr_text_present(img)
            watermark           = _cv_watermark_check(img, use_blip_flag, blip_proc, blip_model)
            drawing             = _drawing_check(img)

            if drawing:
                remark = "Technical / CAD drawing"
            elif watermark:
                remark = "Watermark detected"
            elif blurry:
                remark = "Blurry image"
            elif text_present:
                remark = "Text detected in image"
            else:
                remark = "Valid product image"

            return {
                "url":            url,
                "fetch_status":   "OK",
                "blur_score":     blur_score,
                "blurry":         "Yes" if blurry    else "No",
                "text_detected":  "Yes" if text_present else "No",
                "watermark":      "Yes" if watermark  else "No",
                "autocad_drawing":"Yes" if drawing    else "No",
                "ai_remark":      remark,
            }

        if file3 and st.button("▶ Run Advanced AI QC", key="r3_btn"):

            # Load BLIP only if enabled
            blip_proc_obj  = None
            blip_model_obj = None
            if use_blip:
                blip_proc_obj, blip_model_obj = _load_blip()

            xl           = pd.ExcelFile(file3)
            output_buf   = BytesIO()
            total_sheets = len(xl.sheet_names)
            sheet_prog   = st.progress(0)
            global_start = time.time()

            with pd.ExcelWriter(output_buf, engine="openpyxl") as writer:
                for s_idx, sheet in enumerate(xl.sheet_names):
                    df = xl.parse(sheet)
                    if "image_url" not in df.columns:
                        st.warning(f"Sheet '{sheet}' has no 'image_url' column — skipped.")
                        continue

                    urls_list  = df["image_url"].dropna().tolist()
                    total_rows = len(urls_list)
                    results    = []
                    row_prog   = st.progress(0)
                    row_status = st.empty()
                    s_start    = time.time()

                    # NOTE: BLIP is NOT thread-safe; run sequentially when enabled.
                    # When BLIP is off, use ThreadPoolExecutor for speed.
                    if use_blip:
                        for r_idx, url in enumerate(urls_list):
                            res = _analyze_single(
                                url, BLUR_THRESHOLD_QC,
                                use_blip, blip_proc_obj, blip_model_obj
                            )
                            results.append(res)
                            done    = r_idx + 1
                            elapsed = time.time() - s_start
                            eta     = (elapsed / done) * (total_rows - done) if done > 0 else 0
                            row_prog.progress(done / total_rows)
                            row_status.info(
                                f"Sheet [{sheet}] — {done}/{total_rows} | "
                                f"Elapsed: {fmt_time(elapsed)} | ETA: {fmt_time(eta)}"
                            )
                    else:
                        from concurrent.futures import ThreadPoolExecutor, as_completed as _asc
                        with ThreadPoolExecutor(max_workers=workers3) as ex:
                            fut_map = {
                                ex.submit(
                                    _analyze_single, url,
                                    BLUR_THRESHOLD_QC, False, None, None
                                ): url
                                for url in urls_list
                            }
                            for f in _asc(fut_map):
                                results.append(f.result())
                                done    = len(results)
                                elapsed = time.time() - s_start
                                eta     = (elapsed / done) * (total_rows - done) if done > 0 else 0
                                row_prog.progress(done / total_rows)
                                row_status.info(
                                    f"Sheet [{sheet}] — {done}/{total_rows} | "
                                    f"Elapsed: {fmt_time(elapsed)} | ETA: {fmt_time(eta)}"
                                )

                    row_status.empty()
                    out_df = pd.DataFrame(results, columns=[
                        "url", "fetch_status", "blur_score", "blurry",
                        "text_detected", "watermark", "autocad_drawing", "ai_remark"
                    ])
                    out_df.to_excel(writer, sheet_name=f"{sheet}_Result"[:31], index=False)
                    sheet_prog.progress((s_idx + 1) / total_sheets)

            total_elapsed = time.time() - global_start
            st.success(f"✅ Advanced AI QC Completed in {fmt_time(total_elapsed)}")
            st.download_button(
                "⬇ Download Excel",
                output_buf.getvalue(),
                "advanced_ai_qc.xlsx"
            )


    # ── TAB 4 — SRGAN IMAGE ENHANCER ─────────────────────
    with tab4:

        st.markdown("### ✨ SRGAN Image Enhancer")
        st.info(
            "Upload your SRGAN `.pt` / `.pth` model file and an Excel file with image URLs. "
            "The pipeline downloads each image, runs AI super-resolution, and returns a ZIP "
            "containing enhanced images + a results Excel."
        )

        # ── File uploads — always manual, no disk auto-load ─────────────────
        up1, up2 = st.columns(2)
        with up1:
            srgan_model_file = st.file_uploader(
                "🤖 Upload SRGAN Model (.pt / .pth)",
                type=["pt","pth"], key="srgan_model_upload",
                help="Your trained SRGAN generator weights file"
            )
            if srgan_model_file:
                st.success(f"✅ Model: **{srgan_model_file.name}** "
                           f"({round(srgan_model_file.size/1e6,1)} MB)")
        with up2:
            srgan_excel = st.file_uploader(
                "📊 Upload Excel with image URLs",
                type=["xlsx"], key="srgan_excel",
                help="Must contain a column of image URLs"
            )
            if srgan_excel:
                st.success(f"✅ Excel: **{srgan_excel.name}**")

        # ── Settings ─────────────────────────────────────────────────────────
        s4c1, s4c2 = st.columns(2)
        with s4c1:
            srgan_url_col = st.text_input(
                "URL column name in your Excel",
                value="image_url", key="srgan_url_col",
                help="Exact column header containing the image URLs"
            )
        with s4c2:
            srgan_res_num = st.number_input(
                "Residual blocks (must match training)",
                min_value=1, max_value=32, value=16, step=1, key="srgan_res",
                help="Leave at 16 — verified against your SRGAN.pt"
            )

        # ── Helpers ─────────────────────────────────────────────────────────
        def _srgan_get_ext(url, response):
            parsed = urlparse(url)
            ext    = Path(parsed.path).suffix.lower()
            if ext in [".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".webp"]:
                return ext
            ct = response.headers.get("Content-Type", "")
            mapping = {
                "image/jpeg": ".jpg", "image/png": ".png", "image/bmp": ".bmp",
                "image/tiff": ".tiff", "image/webp": ".webp",
            }
            return mapping.get(ct.split(";")[0].strip(), ".jpg")

        # ── Self-contained SRGAN Generator matching SRGAN.pt exactly ───────────
        import torch.nn as _nn

        class _W(_nn.Module):
            """Wraps layers in .body Sequential to match .pt key naming."""
            def __init__(self, *layers):
                super().__init__()
                self.body = _nn.Sequential(*layers)
            def forward(self, x): return self.body(x)

        class _ResBlock(_nn.Module):
            def __init__(self, ch=64):
                super().__init__()
                self.body = _nn.Sequential(
                    _W(_nn.Conv2d(ch,ch,3,1,1), _nn.BatchNorm2d(ch), _nn.PReLU()),
                    _W(_nn.Conv2d(ch,ch,3,1,1), _nn.BatchNorm2d(ch)),
                )
            def forward(self, x): return x + self.body(x)

        class _UpBlock(_nn.Module):
            def __init__(self, ch=64, scale=2):
                super().__init__()
                self.body = _nn.Sequential(
                    _W(_nn.Conv2d(ch, ch*scale*scale, 3,1,1)),
                    _nn.PixelShuffle(scale),
                    _nn.PReLU(),
                )
            def forward(self, x): return self.body(x)

        class _SRGANGenerator(_nn.Module):
            """
            Exact architecture matching SRGAN.pt — verified zero missing/unexpected keys.
            conv01(9x9) → 16×ResBlock → conv02+BN → 2×UpBlock(×4 total) → last_conv
            Returns (output, None) for compatibility.
            """
            def __init__(self, img_feat=3, n_feats=64, kernel_size=3, num_block=16):
                super().__init__()
                self.conv01    = _W(_nn.Conv2d(img_feat,n_feats,9,1,4), _nn.PReLU())
                self.body      = _nn.Sequential(*[_ResBlock(n_feats) for _ in range(num_block)])
                self.conv02    = _W(_nn.Conv2d(n_feats,n_feats,kernel_size,1,1,bias=True),
                                    _nn.BatchNorm2d(n_feats))
                self.tail      = _nn.Sequential(_UpBlock(n_feats,2), _UpBlock(n_feats,2))
                self.last_conv = _W(_nn.Conv2d(n_feats,img_feat,kernel_size,1,1))
            def forward(self, x):
                h = self.conv01(x)
                h = self.conv02(self.body(h)) + h
                return self.last_conv(self.tail(h)), None

        @st.cache_resource(show_spinner="Loading SRGAN model…")
        def _load_srgan_generator(model_bytes, res_n):
            """Load uploaded .pt bytes — no external srgan_model.py needed."""
            import tempfile as _tf
            device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
            gen    = _SRGANGenerator(img_feat=3, n_feats=64, kernel_size=3, num_block=res_n)
            with _tf.NamedTemporaryFile(suffix=".pt", delete=False) as _tmp:
                _tmp.write(model_bytes); tmp_path = _tmp.name
            state = torch.load(tmp_path, map_location="cpu", weights_only=False)
            if isinstance(state, dict):
                state = state.get("generator", state.get("state_dict", state))
            gen.load_state_dict(state, strict=True)
            gen = gen.to(device); gen.eval()
            return gen, device

        def _srgan_enhance(pil_img, generator, device):
            from torchvision import transforms
            t = transforms.ToTensor()(pil_img)
            t = (t - 0.5) / 0.5
            with torch.no_grad():
                result = generator(t.unsqueeze(0).to(device))
            out = result[0] if isinstance(result, (tuple,list)) else result
            out = out[0].cpu().numpy()
            out = (np.clip(out, -1.0, 1.0) + 1.0) / 2.0
            out = out.transpose(1,2,0)
            from PIL import Image as _PILImg
            return _PILImg.fromarray((out*255.0).astype(np.uint8))

        def _srgan_build_zip(enhanced_dict, excel_bytes):
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr("enhanced_results.xlsx", excel_bytes)
                for name, img_bytes in enhanced_dict.items():
                    zf.writestr(f"enhanced_images/{name}", img_bytes)
            return buf.getvalue()

        # ── Session state for SRGAN results ─────────────────────────────────
        if "srgan_result_zip"     not in st.session_state:
            st.session_state.srgan_result_zip     = None
        if "srgan_result_summary" not in st.session_state:
            st.session_state.srgan_result_summary = None

        # ── Run button ───────────────────────────────────────────────────────
        srgan_run = st.button(
            "▶ Run Enhancement Pipeline", key="srgan_run_btn",
            disabled=(srgan_model_file is None or srgan_excel is None)
        )
        if srgan_model_file is None or srgan_excel is None:
            st.caption("ℹ️ Upload both the model (.pt) and Excel file above to enable.")

        if srgan_run:
            if srgan_excel is None or srgan_model_file is None:
                st.error("Please upload both the model file and Excel file."); st.stop()

            srgan_df = pd.read_excel(srgan_excel, dtype=str)
            if srgan_url_col not in srgan_df.columns:
                st.error(f"Column '{srgan_url_col}' not found. "
                         f"Available: {list(srgan_df.columns)}"); st.stop()

            total_srgan = len(srgan_df)

            with st.spinner("Loading SRGAN model…"):
                try:
                    srgan_model_file.seek(0)
                    model_bytes = srgan_model_file.read()
                    srgan_gen, srgan_device = _load_srgan_generator(
                        model_bytes, int(srgan_res_num)
                    )
                    st.success(f"✅ Model loaded — running on **{srgan_device}**")
                except Exception as e:
                    st.error(f"Failed to load model: {e}"); st.stop()

            # Progress UI
            s4_col1, s4_col2, s4_col3, s4_col4 = st.columns(4)
            ph_total = s4_col1.empty(); ph_enh = s4_col2.empty()
            ph_fail  = s4_col3.empty(); ph_skip = s4_col4.empty()

            def _render_srgan_stat(ph, num, label, color):
                ph.markdown(
                    f'<div style="background:#fff;border:1px solid #DDE2EA;border-radius:12px;padding:1rem;text-align:center;">' +
                    f'<div style="font-family:monospace;font-size:2rem;font-weight:600;color:{color}">{num}</div>' +
                    f'<div style="color:#6B7280;font-size:0.73rem;font-weight:600;text-transform:uppercase;letter-spacing:0.08em;margin-top:0.4rem;">{label}</div>' +
                    f'</div>', unsafe_allow_html=True
                )

            _render_srgan_stat(ph_total, total_srgan, "Total Rows",  "#111827")
            _render_srgan_stat(ph_enh,   0,           "Enhanced",    "#16A34A")
            _render_srgan_stat(ph_fail,  0,           "Failed",      "#DC2626")
            _render_srgan_stat(ph_skip,  0,           "Skipped",     "#6B7280")

            srgan_prog = st.progress(0)
            log_ph     = st.empty()
            log_lines  = []
            srgan_start = time.time()

            def _srgan_log(msg, kind=""):
                css_map = {"ok":"color:#16A34A", "err":"color:#DC2626",
                           "skip":"color:#6B7280", "step":"color:#2563EB;font-weight:600"}
                css = css_map.get(kind, "")
                log_lines.append(f'<span style="{css}">{msg}</span>')
                elapsed = time.time() - srgan_start
                log_ph.markdown(
                    f'<div style="background:#F8FAFC;border:1px solid #DDE2EA;border-radius:10px;' +
                    f'padding:1rem 1.2rem;font-family:monospace;font-size:0.78rem;color:#374151;' +
                    f'max-height:260px;overflow-y:auto;line-height:1.9;">' +
                    "<br>".join(log_lines[-35:]) +
                    f'</div>',
                    unsafe_allow_html=True
                )

            # Step 1 — Download
            _srgan_log("━━━  STEP 1: Downloading images  ━━━", "step")
            downloaded   = {}
            unique_names = [""] * total_srgan
            statuses     = [""] * total_srgan
            cnt_skip = cnt_fail = 0

            for i, url in enumerate(srgan_df[srgan_url_col]):
                srgan_prog.progress((i + 1) / (total_srgan * 2))
                elapsed = time.time() - srgan_start
                eta     = (elapsed / (i + 1)) * (total_srgan - i - 1) if i > 0 else 0

                if pd.isna(url) or str(url).strip() == "":
                    statuses[i] = "Skipped - empty URL"
                    cnt_skip   += 1
                    _srgan_log(f"[{i+1}/{total_srgan}] Skipped — empty URL | ETA: {fmt_time(eta)}", "skip")
                    _render_srgan_stat(ph_skip, cnt_skip, "Skipped", "#6B7280")
                    continue

                url = str(url).strip()
                try:
                    resp             = requests.get(url, timeout=15)
                    resp.raise_for_status()
                    ext              = _srgan_get_ext(url, resp)
                    name             = f"{uuid.uuid4().hex}{ext}"
                    from PIL import Image as _PILImg
                    img              = _PILImg.open(io.BytesIO(resp.content)).convert("RGB")
                    downloaded[i]    = img
                    unique_names[i]  = name
                    statuses[i]      = "downloaded"
                    _srgan_log(f"[{i+1}/{total_srgan}] ✓ {name} | Elapsed: {fmt_time(elapsed)} | ETA: {fmt_time(eta)}", "ok")
                except Exception as e:
                    statuses[i] = f"Failed - {e}"
                    cnt_fail   += 1
                    _srgan_log(f"[{i+1}/{total_srgan}] ✗ {url[:60]}… → {e}", "err")
                    _render_srgan_stat(ph_fail, cnt_fail, "Failed", "#DC2626")

            # Step 2 — Enhance
            valid_imgs = list(downloaded.items())
            _srgan_log(f"━━━  STEP 2: Enhancing {len(valid_imgs)} images  ━━━", "step")
            enhanced_imgs = {}
            enhanced_names = [""] * total_srgan
            cnt_ok = cnt_enh_fail = 0

            for count, (idx, pil_img) in enumerate(valid_imgs):
                srgan_prog.progress(0.5 + (count + 1) / (len(valid_imgs) * 2))
                elapsed = time.time() - srgan_start
                eta     = (elapsed / (count + 1)) * (len(valid_imgs) - count - 1) if count > 0 else 0
                try:
                    enh_img  = _srgan_enhance(pil_img, srgan_gen, srgan_device)
                    enh_name = f"enhanced_{unique_names[idx]}"
                    buf_img  = io.BytesIO()
                    fmt_img  = Path(enh_name).suffix.lstrip(".").upper()
                    fmt_img  = "JPEG" if fmt_img in ("JPG", "") else fmt_img
                    enh_img.save(buf_img, format=fmt_img)
                    enhanced_imgs[enh_name] = buf_img.getvalue()
                    enhanced_names[idx]     = enh_name
                    statuses[idx]           = "Downloaded & Enhanced"
                    cnt_ok += 1
                    _srgan_log(
                        f"[{count+1}/{len(valid_imgs)}] ✨ {enh_name} | "
                        f"Elapsed: {fmt_time(elapsed)} | ETA: {fmt_time(eta)}", "ok"
                    )
                    _render_srgan_stat(ph_enh, cnt_ok, "Enhanced", "#16A34A")
                except Exception as e:
                    statuses[idx] = "Downloaded (enhancement failed)"
                    cnt_enh_fail += 1; cnt_fail += 1
                    _srgan_log(f"[{count+1}/{len(valid_imgs)}] ✗ Enhancement failed — {e}", "err")
                    _render_srgan_stat(ph_fail, cnt_fail, "Failed", "#DC2626")

            srgan_prog.progress(1.0)
            total_elapsed_srgan = time.time() - srgan_start
            _srgan_log(f"━━━  Pipeline complete in {fmt_time(total_elapsed_srgan)}!  ━━━", "step")

            # Build output Excel
            out_srgan_df = pd.DataFrame({
                srgan_url_col:         srgan_df[srgan_url_col],
                "unique_image_name":   unique_names,
                "enhanced_image_name": enhanced_names,
                "status":              statuses,
            })
            excel_buf_srgan = io.BytesIO()
            with pd.ExcelWriter(excel_buf_srgan, engine="openpyxl") as writer:
                out_srgan_df.to_excel(writer, index=False, sheet_name="Results")

            # Pack ZIP and store in session state
            st.session_state.srgan_result_zip = _srgan_build_zip(
                enhanced_imgs, excel_buf_srgan.getvalue()
            )
            st.session_state.srgan_result_summary = {
                "enhanced": cnt_ok, "failed": cnt_fail,
                "skipped": cnt_skip, "total": total_srgan,
            }

        # ── Download (survives reruns) ────────────────────────────────────────
        if st.session_state.srgan_result_zip is not None:
            s = st.session_state.srgan_result_summary
            st.success(
                f"✅ Pipeline complete — Enhanced: {s['enhanced']}  |  "
                f"Failed: {s['failed']}  |  Skipped: {s['skipped']}  |  Total: {s['total']}"
            )
            st.download_button(
                label="⬇️ Download Everything (Excel + Images ZIP)",
                data=st.session_state.srgan_result_zip,
                file_name="srgan_output.zip",
                mime="application/zip",
                use_container_width=True,
            )


# =========================================================
# ─────────────────────────────────────────────────────────
# TOOL 3 — PRODUCT DATA QC
# ─────────────────────────────────────────────────────────
# =========================================================
elif tool == "📊 Product Data QC":

    st.subheader("📊 Product Data Internal QC Checker")

    def clean_str(val):
        if pd.isna(val): return ""
        return str(val).strip()

    def has_extra_spaces(val):
        s = str(val)
        return s != s.strip() or "  " in s

    def mpn_check(mpn, all_mpns, cross_refs):
        if not mpn: return ""
        remark = ""
        if " " in mpn:              remark += "; Contains space"
        if not mpn.isalpha():       remark += "; Non-text"
        if mpn in all_mpns:         remark += "; Duplicate"
        if mpn in cross_refs:       remark += "; Duplicate with CrossRef"
        return remark

    def crossref_check(cross, all_cross, mpns):
        if not cross: return ""
        remark = ""
        if cross in all_cross: remark += "; Duplicate"
        if cross in mpns:      remark += "; Duplicate with MPN"
        return remark

    def title_checks(title, mpn, brand, cross_ref, all_titles):
        if not title: return ""
        remark = []
        t_lower = title.lower()
        if mpn       and mpn.lower()       not in t_lower: remark.append("MPN missing in title")
        if brand     and brand.lower()     not in t_lower: remark.append("Brand missing in title")
        if cross_ref and cross_ref.lower() not in t_lower: remark.append("CrossRef missing in title")
        words   = re.findall(r'\b\w+\b', title.lower())
        counts  = Counter(words)
        repeated = [f"{w}({c})" for w, c in counts.items() if c > 1]
        if repeated: remark.append("Repeated words: " + ", ".join(repeated))
        if title in all_titles: remark.append("Duplicate title")
        if re.search(r"[,.!?;:](?!\s)", title): remark.append("Missing space after punctuation")
        if re.search(r"[^A-Za-z0-9\s\-/.,]", title): remark.append("Unnecessary punctuation")
        if re.search(r"[™®©]", title): remark.append("Contains trademark")
        return "; ".join(remark)

    def numeric_check(val, max_val=70):
        if val is None or str(val).strip() == "": return ""
        try:
            num = float(val)
            if num == 0 or num > max_val: return "; Invalid value"
            return ""
        except:
            return "; Non-numeric"

    def singular_check(val):
        if val and val.endswith("s"): return "; Should be singular"
        return ""

    def plural_check(val):
        if val and not val.endswith("s"): return "; Should be plural"
        return ""

    def upc_check(val):
        if not val: return ""
        clean_upc = val.replace("rp_", "")
        if not clean_upc.isdigit(): return "; Non-numeric UPC"
        return f"; Length({len(clean_upc)})" if not 12 <= len(clean_upc) <= 14 else ""

    def qc_check(df_full, chunk_size=50000):
        all_mpns_global   = set(df_full["Mpn"].dropna().astype(str).str.strip())
        all_cross_global  = set(df_full["Cross Reference"].dropna().astype(str).str.strip())
        all_titles_global = set(df_full["Product_Title"].dropna().astype(str).str.strip())
        results = []
        for start in range(0, len(df_full), chunk_size):
            df = df_full.iloc[start:start+chunk_size].copy()
            new_df = df.copy()
            remark_cols = ["Mpn","Cross Reference","UPC","Product_Title",
                           "Shipping length(inch)","Shipping height(inch)",
                           "Shipping width(inch)","weight(lb)",
                           "Item Name","Parent_name","Brand","brand_url","Entity Id","L3 Name"]
            for col in remark_cols:
                if col in new_df.columns:
                    new_df[col + "_Remark"] = ""
            mpns      = df["Mpn"].dropna().astype(str).str.strip().tolist()
            cross_refs = df["Cross Reference"].dropna().astype(str).str.strip().tolist()
            titles    = df["Product_Title"].dropna().astype(str).str.strip().tolist()
            for idx, row in new_df.iterrows():
                if clean_str(row.get("Mpn")):
                    new_df.at[idx, "Mpn_Remark"] += mpn_check(clean_str(row["Mpn"]), mpns, cross_refs)
                if clean_str(row.get("Cross Reference")):
                    new_df.at[idx, "Cross Reference_Remark"] += crossref_check(clean_str(row["Cross Reference"]), cross_refs, mpns)
                if clean_str(row.get("Product_Title")):
                    new_df.at[idx, "Product_Title_Remark"] = title_checks(
                        clean_str(row["Product_Title"]),
                        clean_str(row.get("Mpn")),
                        clean_str(row.get("Brand")),
                        clean_str(row.get("Cross Reference")),
                        titles
                    )
            for col in ["Shipping length(inch)","Shipping height(inch)","Shipping width(inch)","weight(lb)"]:
                if col in new_df.columns:
                    new_df[col + "_Remark"] += new_df[col].apply(numeric_check)
            if "UPC" in new_df.columns:
                new_df["UPC_Remark"] += new_df["UPC"].astype(str).apply(upc_check)
            if "Item Name" in new_df.columns:
                new_df["Item Name_Remark"] += new_df["Item Name"].astype(str).apply(singular_check)
            if "Parent_name" in new_df.columns:
                new_df["Parent_name_Remark"] += new_df["Parent_name"].astype(str).apply(plural_check)
            for col in remark_cols:
                if col in new_df.columns:
                    new_df[col + "_Remark"] += new_df[col].apply(
                        lambda x: "; Extra spaces" if clean_str(x) and has_extra_spaces(x) else ""
                    )
            results.append(new_df)
        return pd.concat(results, ignore_index=True)

    uploaded_file = st.file_uploader("📤 Upload Excel File", type=["xlsx"])
    if uploaded_file and st.button("🚀 Run QC Check"):
        qc_start = time.time()
        df = pd.read_excel(uploaded_file, dtype=str)
        status_ph = st.empty()
        status_ph.info(f"⏳ Running QC on {len(df)} rows…")
        with st.spinner("Processing QC checks…"):
            final_df = qc_check(df)
            tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            final_df.to_excel(tmp_file.name, index=False, engine="openpyxl")
        elapsed = time.time() - qc_start
        status_ph.empty()
        st.success(f"✅ QC completed in {fmt_time(elapsed)}")
        st.download_button(
            label="⬇️ Download QC Output",
            data=open(tmp_file.name, "rb"),
            file_name="qc_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.subheader("🔍 Preview (First 20 rows)")
        st.dataframe(final_df.head(20))


# =========================================================
# ─────────────────────────────────────────────────────────
# TOOL 4 — A/B TESTING QC (ADMIN)
# ─────────────────────────────────────────────────────────
# =========================================================
elif tool == "🔐 A/B Testing QC (Admin)":

    admin_gate()
    st.subheader("📊 A/B QC Tool — Data Sent vs Live Data (Full)")

    COL_MAP = {
        "Brand":"brand_name","Country of Origin":"country_of_origin",
        "Cross Reference":"cross_ref","Datasheet":"datasheet",
        "Entity Id":"entity_id","Grainger Sku":"grainger_sku",
        "Image Name":"image_name","Item Name":"item_name",
        "L3 Name":"L3_entity_name","Label":"label","Mpn":"mpn",
        "Product_Title":"title","Shipping height(inch)":"ship_height",
        "Shipping length(inch)":"ship_length","Shipping width(inch)":"ship_width",
        "UPC":"gtin","weight(lb)":"weight"
    }

    def as_text(x):
        if pd.isna(x): return ""
        s = str(x).replace("\xa0"," ").replace("\t"," ").replace("\n"," ")
        return " ".join(s.split()).strip()

    def is_empty_after_clean(x):
        return as_text(x) == ""

    def to_num(x):
        try:
            return float(pd.to_numeric(str(x).replace(",",""), errors="coerce"))
        except Exception:
            return np.nan

    def norm_string(ds, ld):     return as_text(ds), as_text(ld)
    def norm_upc(ds, ld):        return as_text(ds).lower().replace("rp_",""), as_text(ld).lower().replace("gtn:","")
    def norm_grainger(ds, ld):   return as_text(ds).lower().replace("rp_",""), as_text(ld).lower().replace("gsku:","")
    def norm_cross_ref(ds, ld):  return as_text(ds).lower().replace("rp_",""), as_text(ld).lower().replace("cross_ref:","")
    def norm_image(ds, ld):
        ds, ld = as_text(ds), as_text(ld)
        if ds.lower() == "/notavail.webp": ds = ""
        if ld.lower() == "/notavail.webp": ld = ""
        return ds, ld

    def case_insensitive_updated(ds, ld):
        ds_n, ld_n = as_text(ds), as_text(ld)
        if ld_n == "":                   return "Not Updated"
        if ds_n.lower() == ld_n.lower(): return "Updated"
        return "Mismatch"

    def num_with_tolerance(ds, ld, tol=0.01):
        ds_n, ld_n = to_num(ds), to_num(ld)
        if pd.isna(ld_n):                    return "Not Updated"
        if abs(ds_n - ld_n) <= tol:         return "Updated"
        return "Mismatch"

    def weights_compare(ds_lbs, ld_kg):
        lbs  = to_num(ds_lbs)
        ldkg = to_num(ld_kg)
        if pd.isna(lbs): return "", ""
        ds_kg = round(lbs * 0.45, 2)
        if pd.isna(ldkg): return str(ds_kg), ""
        return str(ds_kg), str(round(ldkg, 2))

    col1, col2 = st.columns(2)
    with col1:
        file1 = st.file_uploader("Data Sent (file1.xlsx)", type=["xlsx"])
    with col2:
        file2 = st.file_uploader("Live Data (file2.xlsx)", type=["xlsx"])

    if st.button("🔍 Run QC"):
        if not file1 or not file2:
            st.error("Upload both files"); st.stop()

        ab_start = time.time()
        data_sent = pd.read_excel(file1)
        live_data = pd.read_excel(file2)

        df = data_sent.merge(live_data, left_on="Entity Id", right_on="entity_id", how="left", suffixes=("_ds","_ld"))

        def get_entity_status(row):
            disc = str(row.get("Discontinue/Remove","")).strip().lower()
            if pd.isna(row["entity_id"]):
                return "removed" if disc in ["remove","discontinue"] else "Missing in Live Data"
            else:
                return "not removed" if disc in ["remove","discontinue"] else "Present in Live Data"

        df["Entity Id_status"] = df.apply(get_entity_status, axis=1)
        status_cols = []

        for ds_col, ld_col in COL_MAP.items():
            if ds_col == "Entity Id": continue
            out_status = []
            for _, row in df.iterrows():
                ds_v, ld_v = row.get(ds_col,""), row.get(ld_col,"")
                if str(row.get("Discontinue/Remove","")).strip().lower() in ["remove","discontinue"]:
                    out_status.append(""); continue
                if is_empty_after_clean(ds_v):
                    out_status.append(""); continue
                if str(ds_v).strip().lower() == "delete" and is_empty_after_clean(ld_v):
                    out_status.append("Updated"); continue
                if ds_col in ["Product_Title","L3 Name","Item Name"]:
                    out_status.append(case_insensitive_updated(ds_v, ld_v))
                elif ds_col in ["Shipping height(inch)","Shipping length(inch)","Shipping width(inch)"]:
                    out_status.append(num_with_tolerance(ds_v, ld_v, 0.01))
                elif ds_col == "weight(lb)":
                    ds_kg, ld_kg = weights_compare(ds_v, ld_v)
                    if ld_kg == "":               out_status.append("Not Updated")
                    elif abs(float(ds_kg)-float(ld_kg)) <= 0.01: out_status.append("Updated")
                    else:                         out_status.append("Mismatch")
                elif ds_col == "Cross Reference":
                    ds_n, ld_n = norm_cross_ref(ds_v, ld_v)
                    out_status.append("Updated" if ds_n==ld_n else "Mismatch")
                elif ds_col == "UPC":
                    ds_n, ld_n = norm_upc(ds_v, ld_v)
                    out_status.append("Updated" if ds_n==ld_n else "Mismatch")
                elif ds_col == "Grainger Sku":
                    ds_n, ld_n = norm_grainger(ds_v, ld_v)
                    out_status.append("Updated" if ds_n==ld_n else "Mismatch")
                elif ds_col == "Image Name":
                    ds_n, ld_n = norm_image(ds_v, ld_v)
                    out_status.append("Updated" if ld_n else "Not Updated")
                else:
                    ds_n, ld_n = norm_string(ds_v, ld_v)
                    out_status.append("Updated" if ds_n==ld_n else "Mismatch")
            df[f"{ds_col}_status"] = out_status
            status_cols.append(f"{ds_col}_status")

        report_rows = []
        for _, row in df.iterrows():
            if row["Entity Id_status"] == "not removed":
                report_rows.append({"Entity Id":row["Entity Id"],"Problem Columns":"Discontinue/Remove","Mismatch Details":"Not removed / Not discontinued"})
                continue
            probs, details = [], []
            for ds_col in COL_MAP:
                if ds_col == "Entity Id": continue
                st_val = row.get(f"{ds_col}_status","")
                if st_val in ["Mismatch","Not Updated"]:
                    probs.append(ds_col); details.append(ds_col)
            if probs:
                report_rows.append({"Entity Id":row["Entity Id"],"Problem Columns":", ".join(probs),"Mismatch Details":", ".join(details)})

        elapsed = time.time() - ab_start
        report_df = pd.DataFrame(report_rows)
        st.success(f"✅ QC Completed in {fmt_time(elapsed)}")
        st.dataframe(df.head(100))
        st.subheader("QC Report")
        st.dataframe(report_df)


# =========================================================
# ─────────────────────────────────────────────────────────
# TOOL 5 — MRO PRICE SHEET MAPPER  (price1.py)
# ─────────────────────────────────────────────────────────
# =========================================================
elif tool == "💰 MRO Price Sheet Mapper":

    st.header("💰 MRO Price Sheet Mapper")

    # ── Helper ────────────────────────────────────────────
    def clean_mpn(val, prefix=""):
        if pd.isna(val): return ""
        s = str(val).replace(prefix, "") if prefix else str(val)
        return re.sub(r"[^a-zA-Z0-9]", "", s).upper()

    BLANK = "(blank — skip)"

    def selectbox_with_blank(label, cols, default="", key=None):
        options = [BLANK] + list(cols)
        idx     = options.index(default) if default in options else 0
        return st.selectbox(label, options, index=idx, key=key)

    # ── 1. File Upload ────────────────────────────────────
    st.subheader("1. Upload Files")
    pc1, pc2 = st.columns(2)
    with pc1:
        my_file  = st.file_uploader("Your Data File",              type=["xlsx","xls"], key="pm_my")
    with pc2:
        mfr_file = st.file_uploader("Manufacturer Price Sheet",    type=["xlsx","xls"], key="pm_mfr")

    if not my_file or not mfr_file:
        st.info("Please upload both files to continue.")
        st.stop()

    # ── Load headers ──────────────────────────────────────
    @st.cache_data
    def get_cols(file, sheet=0):
        return list(pd.read_excel(file, sheet_name=sheet, nrows=0).columns)

    my_cols  = get_cols(my_file)
    mfr_cols = get_cols(mfr_file)

    # ── 2. Available columns ──────────────────────────────
    st.subheader("2. Available Columns")
    pc1, pc2 = st.columns(2)
    with pc1:
        st.markdown("**Your Data File Columns:**")
        st.code("  |  ".join(my_cols), language=None)
    with pc2:
        st.markdown("**Manufacturer Sheet Columns:**")
        st.code("  |  ".join(mfr_cols), language=None)

    # ── 3. Manufacturer mapping ───────────────────────────
    st.subheader("3. Manufacturer Sheet — Column Mapping")
    pc1, pc2, pc3 = st.columns(3)
    with pc1:
        mfr_mpn_col = st.selectbox("MPN Column", mfr_cols, index=mfr_cols.index("Reference") if "Reference" in mfr_cols else 0, key="pm_mfr_mpn")
    with pc2:
        mfr_price_col = st.selectbox("List Price Column", mfr_cols, index=mfr_cols.index("€ Exc. VAT") if "€ Exc. VAT" in mfr_cols else 0, key="pm_mfr_price")
    with pc3:
        mfr_base_price_col_sel = selectbox_with_blank("Base Price Column (blank = same as List Price)", mfr_cols, BLANK, key="pm_mfr_base")
        mfr_base_price_col     = mfr_base_price_col_sel if mfr_base_price_col_sel != BLANK else mfr_price_col

    pc1, pc2 = st.columns(2)
    with pc1:
        mfr_qty_col_sel  = selectbox_with_blank("Quantity Column (blank = skip)",  mfr_cols, BLANK, key="pm_mfr_qty")
        mfr_qty_col      = mfr_qty_col_sel if mfr_qty_col_sel != BLANK else ""
    with pc2:
        mfr_pack_qty_col_sel = selectbox_with_blank("Pack QTY Column (blank = skip)", mfr_cols, BLANK, key="pm_mfr_pack")
        mfr_pack_qty_col     = mfr_pack_qty_col_sel if mfr_pack_qty_col_sel != BLANK else ""

    # ── 4. Your data mapping ──────────────────────────────
    st.subheader("4. Your Data File — Column Mapping")
    pc1, pc2, pc3 = st.columns(3)
    with pc1:
        my_mpn_col    = st.selectbox("MPN Column",     my_cols, index=my_cols.index("mpn")      if "mpn"      in my_cols else 0, key="pm_my_mpn")
        my_mpn_prefix = st.text_input("MPN Prefix to Remove", value="rp_", key="pm_mpn_pfx")
    with pc2:
        my_entity_id_col = st.selectbox("Entity Id Column", my_cols, index=my_cols.index("entity_id")    if "entity_id"    in my_cols else 0, key="pm_eid")
        my_sku_col       = st.selectbox("SKU Column",       my_cols, index=my_cols.index("sku")           if "sku"          in my_cols else 0, key="pm_sku")
        my_sku_prefix    = st.text_input("SKU Prefix to Remove", value="sku:", key="pm_sku_pfx")
    with pc3:
        my_price_source_col = st.selectbox("Price Source Column", my_cols, index=my_cols.index("price_source") if "price_source" in my_cols else 0, key="pm_ps")

    # ── 5. Fixed values ───────────────────────────────────
    st.subheader("5. Fixed Values")
    pc1, pc2, pc3 = st.columns(3)
    with pc1:
        brand_name = st.text_input("Brand Name", value="MASTER LOCK", key="pm_brand")
        brand_url  = st.text_input("Brand URL",  value="b/master-lock", key="pm_burl")
    with pc2:
        margin   = st.text_input("Margin %",   value="15%", key="pm_margin")
        discount = st.text_input("Discount %", value="5%",  key="pm_discount")
    with pc3:
        price_source_assignee = st.text_input("Price Source Assignee", value="DHRUV",                 key="pm_psa")
        price_source          = st.text_input("Price Source",           value="Safety CIAL July 2026", key="pm_ps2")
        oem_price_date        = st.text_input("OEM Price Date",          value="July 2026",             key="pm_opd")

    # ── 6. Generate ───────────────────────────────────────
    st.subheader("6. Generate")

    if st.button("🚀 Generate Price Sheet", type="primary", use_container_width=True, key="pm_gen"):
        pm_start = time.time()
        pm_status = st.empty()

        pm_status.info("⏳ Loading files…")
        my_file.seek(0); mfr_file.seek(0)
        my_data = pd.read_excel(my_file)
        mfr     = pd.read_excel(mfr_file, sheet_name=0)

        pm_status.info("⏳ Processing data…")
        mfr[mfr_price_col] = pd.to_numeric(mfr[mfr_price_col], errors="coerce")
        if mfr_base_price_col != mfr_price_col:
            mfr[mfr_base_price_col] = pd.to_numeric(mfr[mfr_base_price_col], errors="coerce")

        mfr["mpn_clean"]     = mfr[mfr_mpn_col].apply(clean_mpn)
        my_data["mpn_clean"] = my_data[my_mpn_col].apply(lambda x: clean_mpn(x, my_mpn_prefix))

        if my_price_source_col in my_data.columns:
            grainger_mpn_set = set(
                my_data.loc[
                    my_data[my_price_source_col].astype(str).str.strip().str.lower().isin(["grainger"]),
                    "mpn_clean"
                ]
            )
        else:
            grainger_mpn_set = set()

        matched             = mfr.merge(my_data[[my_entity_id_col, my_sku_col, "mpn_clean"]], on="mpn_clean", how="left")
        matched_count       = matched[my_entity_id_col].notna().sum()
        unmatched_count     = matched[my_entity_id_col].isna().sum()
        mfr_mpn_set         = set(mfr["mpn_clean"])
        my_data_only_rows   = my_data[~my_data["mpn_clean"].isin(mfr_mpn_set)].copy()
        call_for_price_count = len(my_data_only_rows)
        mfr_dup_mpns        = set(mfr["mpn_clean"][mfr["mpn_clean"].duplicated(keep=False)])
        my_dup_mpns         = set(my_data["mpn_clean"][my_data["mpn_clean"].duplicated(keep=False)])

        pm_status.info(f"⏳ Building {len(matched)} output rows… Elapsed: {fmt_time(time.time()-pm_start)}")

        output_rows = []
        for i, (_, row) in enumerate(matched.iterrows(), start=1):
            entity_id  = str(int(row[my_entity_id_col])) if pd.notna(row[my_entity_id_col]) else ""
            raw_sku    = str(row[my_sku_col]) if pd.notna(row[my_sku_col]) else ""
            sku        = raw_sku.replace(my_sku_prefix, "").strip() if my_sku_prefix else raw_sku
            list_price = round(row[mfr_price_col], 2) if pd.notna(row[mfr_price_col]) else "Call for Price"
            base_price = round(row[mfr_base_price_col], 2) if pd.notna(row[mfr_base_price_col]) else "Call for Price"
            avail      = str(row.get("Availability","")) if pd.notna(row.get("Availability")) else ""
            mpn_clean  = str(row["mpn_clean"])
            grainger_product = mpn_clean in grainger_mpn_set

            if grainger_product:
                remark = "Grainger Product"
            elif mpn_clean in mfr_dup_mpns and mpn_clean in my_dup_mpns:
                remark = "Duplicate MPN Found (Manufacturer & Data File)"
            elif mpn_clean in mfr_dup_mpns:
                remark = "Duplicate MPN Found (Manufacturer Sheet)"
            elif mpn_clean in my_dup_mpns:
                remark = "Duplicate MPN Found (Data File)"
            elif list_price == "Call for Price":
                remark = "Call for Price (Manufacturer Sheet)"
            else:
                remark = ""

            output_rows.append({
                "S.No":                 i,
                "Entity Id":            entity_id,
                "sku":                  sku,
                "Mpn":                  str(row[mfr_mpn_col]) if pd.notna(row[mfr_mpn_col]) else "",
                "Mpn_clean":            mpn_clean,
                "Brand":                brand_name,
                "Brand_url":            brand_url,
                "Product Url":          "",
                "list_price":           "" if grainger_product else list_price,
                "Discount":             "" if grainger_product else discount,
                "base_price":           "" if grainger_product else base_price,
                "Margin":               "" if grainger_product else margin,
                "Free Shipping Margin": "",
                "Quantity":             (str(row[mfr_qty_col]) if mfr_qty_col and pd.notna(row.get(mfr_qty_col)) else ""),
                "Pack QTY":             (str(row[mfr_pack_qty_col]) if mfr_pack_qty_col and pd.notna(row.get(mfr_pack_qty_col)) else ""),
                "Stock":                avail,
                "Price Source Assignee": price_source_assignee,
                "Price Source":          price_source,
                "OEM Price Date":        oem_price_date,
                "Discontinued/Remove":  "Yes" if avail == "Available until stock depletion" else "No",
                "Match Status":         "Grainger Product" if grainger_product else ("Matched" if entity_id else "Not Found"),
                "Remark":               remark,
            })

        for i2, (_, row) in enumerate(my_data_only_rows.iterrows(), start=len(output_rows)+1):
            raw_sku    = str(row[my_sku_col]) if pd.notna(row[my_sku_col]) else ""
            sku        = raw_sku.replace(my_sku_prefix, "").strip() if my_sku_prefix else raw_sku
            entity_id  = str(int(row[my_entity_id_col])) if pd.notna(row[my_entity_id_col]) else ""
            mpn_clean  = str(row["mpn_clean"])
            remark     = "Duplicate MPN Found (Data File) - Call for Price" if mpn_clean in my_dup_mpns else "Not in Manufacturer Sheet - Call for Price"
            output_rows.append({
                "S.No":                 i2,
                "Entity Id":            entity_id,
                "sku":                  sku,
                "Mpn":                  str(row[my_mpn_col]) if pd.notna(row[my_mpn_col]) else "",
                "Mpn_clean":            mpn_clean,
                "Brand":                brand_name,
                "Brand_url":            brand_url,
                "Product Url":          "",
                "list_price":           "Call for Price",
                "Discount":             "",
                "base_price":           "Call for Price",
                "Margin":               "",
                "Free Shipping Margin": "",
                "Quantity":             "",
                "Pack QTY":             "",
                "Stock":                "",
                "Price Source Assignee": price_source_assignee,
                "Price Source":          "",
                "OEM Price Date":        "",
                "Discontinued/Remove":  "No",
                "Match Status":         "Call for Price",
                "Remark":               remark,
            })

        out_df = pd.DataFrame(output_rows)

        pm_status.info(f"⏳ Creating Excel file… Elapsed: {fmt_time(time.time()-pm_start)}")

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Price Sheet"

        header_fill    = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
        header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        matched_fill   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
        unmatched_fill = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
        duplicate_fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
        callprice_fill = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
        grainger_fill  = PatternFill("solid", start_color="E2CFEE", end_color="E2CFEE")
        thin           = Side(style="thin", color="BFBFBF")
        border         = Border(left=thin, right=thin, top=thin, bottom=thin)
        data_font      = Font(name="Arial", size=9)

        cols = list(out_df.columns)
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row_idx, row_data in enumerate(out_df.itertuples(index=False), start=2):
            remark_val = row_data[-1]
            match_val  = row_data[-2]
            if match_val == "Grainger Product":      row_fill = grainger_fill
            elif match_val == "Call for Price":       row_fill = callprice_fill
            elif remark_val and "Duplicate" in str(remark_val): row_fill = duplicate_fill
            elif match_val == "Matched":              row_fill = matched_fill
            else:                                     row_fill = unmatched_fill
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font; cell.alignment = Alignment(vertical="center")
                cell.border = border;  cell.fill = row_fill

        col_widths = [5,10,12,12,12,14,14,14,11,10,11,10,18,10,10,16,18,20,12,16,12]
        for ci, cw in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = cw
        ws.row_dimensions[1].height = 35
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Summary"
        ws2["A1"].font = Font(name="Arial", bold=True, size=14)
        for r, (label, val) in enumerate([
            ("Total Products",            len(out_df)),
            ("Matched (Entity Id + SKU)", int(matched_count)),
            ("Not Matched",               int(unmatched_count)),
            ("Call for Price",            int(call_for_price_count)),
            ("Grainger Products",         len(out_df[out_df["Match Status"]=="Grainger Product"])),
            ("Match Rate",                f"{round(matched_count/len(out_df)*100,1)}%"),
            ("Brand",                     brand_name),
            ("Margin",                    margin),
            ("Discount",                  discount),
            ("Price Source Assignee",     price_source_assignee),
            ("Price Source",              price_source),
            ("OEM Price Date",            oem_price_date),
        ], start=3):
            ws2[f"A{r}"] = label; ws2[f"B{r}"] = val
            ws2[f"A{r}"].font = Font(name="Arial", size=10)
            ws2[f"B{r}"].font = Font(name="Arial", bold=True, size=10)
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        elapsed = time.time() - pm_start
        pm_status.empty()

        # ── Results ───────────────────────────────────────
        grainger_count = len(out_df[out_df["Match Status"]=="Grainger Product"])
        dup_count      = len(out_df[out_df["Remark"].str.contains("Duplicate", na=False)])

        rc1,rc2,rc3,rc4,rc5,rc6 = st.columns(6)
        rc1.metric("Total",             len(out_df))
        rc2.metric("✅ Matched",        int(matched_count))
        rc3.metric("❌ Not Found",      int(unmatched_count))
        rc4.metric("📞 Call for Price", int(call_for_price_count))
        rc5.metric("🟣 Grainger",       int(grainger_count))
        rc6.metric("🟡 Duplicates",     int(dup_count))

        if len(mfr_dup_mpns) > 0 or len(my_dup_mpns) > 0:
            st.warning(f"⚠️ {len(mfr_dup_mpns | my_dup_mpns)} duplicate MPN(s) detected — marked in Remark column.")

        st.success(f"✅ Price sheet generated in {fmt_time(elapsed)}")
        st.download_button(
            label="⬇️ Download Output_Price_Sheet.xlsx",
            data=buf,
            file_name="Output_Price_Sheet.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        st.subheader("Preview (first 100 rows)")
        st.dataframe(out_df.head(100), use_container_width=True)


# =========================================================
# FOOTER
# =========================================================
st.markdown("---")
st.markdown("**Owned by Raptor Supplies | Created by Shubham Sisodia**")